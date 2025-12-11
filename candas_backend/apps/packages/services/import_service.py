from django.db import transaction
from apps.packages.models import Package, PackageImport


class PackageImportService:
    """Servicio para importar paquetes desde archivos Excel."""
    # Reglas de encabezados para soportar estructura de guias_test.xlsx
    REQUIRED_COLUMNS = ['nro_master', 'guide_number', 'name', 'address', 'city', 'province', 'phone_number']
    # Orden requerido por el usuario
    EXPECTED_HEADER_ORDER = [
        'guide_number',
        'nro_master',
        'notes',
        'city',
        'name',
        'address',
        'phone_number',
        'province',
    ]
    HEADER_ALIASES = {
        'nro master': 'nro_master',
        'nro_master': 'nro_master',
        'master': 'nro_master',
        'guia': 'guide_number',
        'nro guia': 'guide_number',
        'guide_number': 'guide_number',
        'guia_number': 'guide_number',
        'destinatario': 'name',
        'nombre': 'name',
        'name': 'name',
        'direccion': 'address',
        'address': 'address',
        'ciudad': 'city',
        'city': 'city',
        'provincia': 'province',
        'province': 'province',
        'telefono': 'phone_number',
        'celular': 'phone_number',
        'phone': 'phone_number',
        'phone_number': 'phone_number',
        # opcionales
        'notas': 'notes',
        'notes': 'notes',
        'status': 'status',
    }
    
    @staticmethod
    def parse_excel_file(file_path):
        """
        Parsea un archivo Excel y retorna lista de diccionarios con datos de paquetes.
        
        Args:
            file_path (str): Ruta del archivo Excel
            
        Returns:
            tuple: (lista de diccionarios, lista de errores)
        """
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(file_path)
            worksheet = workbook.active
            
            packages = []
            errors = []
            
            # Encabezados originales (fila 1)
            raw_headers = [cell.value for cell in worksheet[1]]
            header_lower = [str(h).strip().lower() if h else '' for h in raw_headers]

            # Normalizar usando alias
            normalized_headers = [PackageImportService.HEADER_ALIASES.get(h, h) for h in header_lower]
            # Eliminar encabezados vacíos al final
            normalized_nonempty = [h for h in normalized_headers if h]

            # Validar orden exacto requerido
            if normalized_nonempty != PackageImportService.EXPECTED_HEADER_ORDER:
                errors.append(
                    "El archivo no cumple el ORDEN requerido de columnas. "
                    "Orden esperado (8 columnas exactas): "
                    + " -> ".join(PackageImportService.EXPECTED_HEADER_ORDER)
                )
                return packages, errors
            
            # Primero, leer todas las filas y agrupar por guide_number
            rows_by_guide = {}
            
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Crear diccionario de datos usando encabezados
                    row_data = {}
                    for idx, header in enumerate(normalized_nonempty):
                        if header and idx < len(row):
                            row_data[str(header)] = row[idx]
                    
                    # Obtener guide_number y validar que no sea None, vacío o "None"
                    guide_value = row_data.get('guide_number')
                    if guide_value is None or str(guide_value).strip() in ['', 'None', 'none', 'NONE']:
                        continue  # Saltar filas sin guía válida
                    
                    guide_number = str(guide_value).strip()
                    
                    # Validación adicional
                    if not guide_number:
                        continue  # Saltar filas sin guía
                    
                    # Si es la primera vez que vemos este guide_number, inicializar
                    if guide_number not in rows_by_guide:
                        rows_by_guide[guide_number] = {
                            'nro_master': str(row_data.get('nro_master', '')).strip(),
                            'guide_number': guide_number,
                            'name': str(row_data.get('name', '')).strip() if row_data.get('name') else '',
                            'addresses': [],
                            'city': str(row_data.get('city', '')).strip() if row_data.get('city') else '',
                            'province': str(row_data.get('province', '')).strip() if row_data.get('province') else '',
                            'phone_number': str(row_data.get('phone_number', '')).strip() if row_data.get('phone_number') else '',
                            'status': str(row_data.get('status', '')).strip() if row_data.get('status') else 'EN_BODEGA',
                            'notes': str(row_data.get('notes', '')).strip() if row_data.get('notes') else '',
                            'row_nums': [row_num]
                        }
                    else:
                        # Agregar información adicional de esta fila
                        rows_by_guide[guide_number]['row_nums'].append(row_num)
                    
                    # Agregar dirección si existe
                    address = str(row_data.get('address', '')).strip()
                    if address:
                        rows_by_guide[guide_number]['addresses'].append(address)
                    
                    # Actualizar campos si estaban vacíos en la primera fila
                    if not rows_by_guide[guide_number]['nro_master'] and row_data.get('nro_master'):
                        rows_by_guide[guide_number]['nro_master'] = str(row_data.get('nro_master', '')).strip()
                    if not rows_by_guide[guide_number]['name'] and row_data.get('name'):
                        rows_by_guide[guide_number]['name'] = str(row_data.get('name', '')).strip()
                    if not rows_by_guide[guide_number]['city'] and row_data.get('city'):
                        rows_by_guide[guide_number]['city'] = str(row_data.get('city', '')).strip()
                    if not rows_by_guide[guide_number]['province'] and row_data.get('province'):
                        rows_by_guide[guide_number]['province'] = str(row_data.get('province', '')).strip()
                    if not rows_by_guide[guide_number]['phone_number'] and row_data.get('phone_number'):
                        rows_by_guide[guide_number]['phone_number'] = str(row_data.get('phone_number', '')).strip()
                    
                except Exception as e:
                    errors.append(f'Fila {row_num}: Error al leer: {str(e)}')
            
            # Convertir los datos agrupados en paquetes
            def normalize(val):
                s = str(val).strip() if val else ''
                return s if s else 'DATO FALTANTE'
            
            for guide_number, data in rows_by_guide.items():
                # Unir todas las direcciones con un espacio
                combined_address = ' '.join(data['addresses']).strip() if data['addresses'] else ''
                
                package_data = {
                    'nro_master': data['nro_master'],
                    'guide_number': guide_number,
                    'name': normalize(data['name']),
                    'address': normalize(combined_address),
                    'city': normalize(data['city']),
                    'province': normalize(data['province']),
                    'phone_number': normalize(data['phone_number']),
                    'status': data['status'] or 'EN_BODEGA',
                    'notes': normalize(data['notes']),
                }
                
                packages.append(package_data)
            
            return packages, errors
            
        except Exception as e:
            return [], [f'Error al leer archivo: {str(e)}']
    
    @staticmethod
    def import_packages(packages_data):
        """
        Importa una lista de paquetes a la base de datos.
        
        Args:
            packages_data (list): Lista de diccionarios con datos de paquetes
            
        Returns:
            tuple: (cantidad creados, cantidad errores, lista de errores)
        """
        created_count = 0
        error_count = 0
        errors = []
        
        def trunc(val: str, max_len: int):
            s = str(val or '')
            return s[:max_len]
        
        def normalize_field(value):
            """Normaliza un campo de texto: elimina espacios extra, saltos de línea, y caracteres especiales"""
            if not value:
                return ''
            # Convertir a string y eliminar espacios al inicio y final
            text = str(value).strip()
            # Eliminar múltiples espacios
            text = ' '.join(text.split())
            # Eliminar caracteres de control y saltos de línea
            text = ''.join(char for char in text if char.isprintable() or char in ['\n', '\t'])
            # Reemplazar múltiples saltos de línea con uno solo
            import re
            text = re.sub(r'\n\s*\n', '\n', text)
            return text

        for idx, pkg_data in enumerate(packages_data, start=1):
            with transaction.atomic():
                try:
                    # Normalizar y sanitizar longitudes según modelo
                    pkg_data['nro_master'] = trunc(normalize_field(pkg_data.get('nro_master', '')), 50)
                    pkg_data['guide_number'] = trunc(normalize_field(pkg_data.get('guide_number', '')), 50)
                    
                    # Validar que guide_number no esté vacío
                    if not pkg_data['guide_number'] or pkg_data['guide_number'] in ['None', 'none', 'NONE']:
                        error_count += 1
                        errors.append(f'Paquete {idx}: guide_number vacío o inválido')
                        continue
                    
                    pkg_data['name'] = trunc(normalize_field(pkg_data.get('name', '')), 100)
                    pkg_data['address'] = trunc(normalize_field(pkg_data.get('address', '')), 200)
                    pkg_data['city'] = trunc(normalize_field(pkg_data.get('city', '')), 100)
                    pkg_data['province'] = trunc(normalize_field(pkg_data.get('province', '')), 100)
                    pkg_data['phone_number'] = trunc(normalize_field(pkg_data.get('phone_number', '')), 20)
                    pkg_data['notes'] = trunc(normalize_field(pkg_data.get('notes', '')), 1000)

                    # Validar unicidad solo de guide_number
                    if Package.objects.filter(guide_number=pkg_data['guide_number']).exists():
                        error_count += 1
                        errors.append(f'Paquete {idx}: guide_number "{pkg_data["guide_number"]}" ya existe')
                        continue

                    Package.objects.create(
                        nro_master=pkg_data['nro_master'],
                        guide_number=pkg_data['guide_number'],
                        name=pkg_data['name'],
                        address=pkg_data['address'],
                        city=pkg_data['city'],
                        province=pkg_data['province'],
                        phone_number=pkg_data['phone_number'],
                        status=pkg_data.get('status', 'EN_BODEGA'),
                        notes=pkg_data.get('notes', '')
                    )
                    created_count += 1
                except Exception as e:
                    error_count += 1
                    errors.append(f'Paquete {idx}: {str(e)}')
        
        return created_count, error_count, errors
