"""
Servicio para exportar paquetes a Excel y PDF
"""
from typing import TYPE_CHECKING
from django.http import HttpResponse
from django.db.models import QuerySet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from io import BytesIO
from datetime import datetime

if TYPE_CHECKING:
    from ..models import Package


class PackageExportService:
    """Servicio para exportar paquetes a diferentes formatos"""
    
    # Campos disponibles para exportación
    AVAILABLE_FIELDS = {
        'guide_number': 'Número de Guía',
        'nro_master': 'Número Master',
        'name': 'Nombre Destinatario',
        'address': 'Dirección',
        'city': 'Ciudad',
        'province': 'Provincia',
        'phone_number': 'Teléfono',
        'status': 'Estado',
        'shipment_type_display': 'Tipo de Envío',
        'transport_agency_name': 'Agencia de Transporte',
        'delivery_agency_name': 'Agencia de Reparto',
        'effective_destiny': 'Destino Efectivo',
        'effective_guide_number': 'Guía Efectiva',
        'pull_name': 'Saca',
        'batch_name': 'Lote',
        'notes': 'Notas',
        'hashtags': 'Hashtags',
        'created_at': 'Fecha Creación',
        'updated_at': 'Fecha Actualización'
    }
    
    @classmethod
    def get_package_field_value(cls, package: "Package", field_name: str) -> str:
        """
        Extrae el valor de un campo del paquete
        Maneja campos relacionados, computados y formatea correctamente
        """
        try:
            match field_name:
                case 'status':
                    return package.get_status_display()
                
                case 'shipment_type_display':
                    return package.get_shipment_type_display()
                
                case 'transport_agency_name':
                    if package.transport_agency:
                        return package.transport_agency.name
                    # Si no tiene agencia propia, obtener la efectiva
                    effective_agency = package.get_shipping_agency()
                    return effective_agency.name if effective_agency else 'N/A'
                
                case 'delivery_agency_name':
                    return package.delivery_agency.name if package.delivery_agency else 'N/A'
                
                case 'effective_destiny':
                    return package.get_effective_destiny() or 'N/A'
                
                case 'effective_guide_number':
                    return package.get_shipping_guide_number() or 'N/A'
                
                case 'pull_name':
                    return f"Saca-{package.pull.id}" if package.pull else 'N/A'
                
                case 'batch_name':
                    if batch := package.get_batch():
                        return f"Lote-{batch.id}"
                    return 'N/A'
                
                case 'created_at' | 'updated_at':
                    if date_value := getattr(package, field_name, None):
                        return date_value.strftime('%Y-%m-%d %H:%M:%S')
                    return 'N/A'
                
                case 'hashtags':
                    if hashtags := getattr(package, field_name, None):
                        return ', '.join(hashtags) if isinstance(hashtags, list) else str(hashtags)
                    return ''
                
                case _:
                    # Campo directo del modelo
                    value = getattr(package, field_name, None)
                    return str(value) if value is not None else ''
                
        except Exception as e:
            print(f"Error al obtener campo {field_name}: {str(e)}")
            return 'Error'
    
    @classmethod
    def generate_excel(cls, queryset: QuerySet["Package"], columns_config: list[str]) -> HttpResponse:
        """
        Genera un archivo Excel con los paquetes especificados
        
        Args:
            queryset: QuerySet de Package
            columns_config: Lista de IDs de campos a incluir en orden
        
        Returns:
            HttpResponse con el archivo Excel
        """
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Paquetes"
        
        # Contar total de paquetes
        total_packages = queryset.count()
        
        # Si no hay columnas especificadas, usar todas
        if not columns_config:
            columns_config = list(cls.AVAILABLE_FIELDS.keys())
        
        # Estilos
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        title_font = Font(bold=True, size=14)
        info_font = Font(size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Obtener letra de la última columna
        last_col_letter = get_column_letter(len(columns_config))
        
        # Agregar encabezado del reporte
        row_num = 1
        ws.merge_cells(f'A{row_num}:{last_col_letter}{row_num}')
        title_cell = ws[f'A{row_num}']
        title_cell.value = f"Reporte de Paquetes - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        row_num += 1
        
        # Agregar información del total
        ws.merge_cells(f'A{row_num}:{last_col_letter}{row_num}')
        info_cell = ws[f'A{row_num}']
        info_cell.value = f"Total de Paquetes: {total_packages}"
        info_cell.font = info_font
        info_cell.alignment = Alignment(horizontal='left', vertical='center')
        row_num += 1
        
        # Línea en blanco
        row_num += 1
        
        # Agregar headers
        headers = [cls.AVAILABLE_FIELDS.get(col, col) for col in columns_config]
        header_row = row_num
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        row_num += 1
        
        # Agregar datos
        for package in queryset:
            for col_idx, field_name in enumerate(columns_config, start=1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.value = cls.get_package_field_value(package, field_name)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            row_num += 1
        
        # Agregar fila de total al final
        total_row = row_num
        ws.merge_cells(f'A{total_row}:{last_col_letter}{total_row}')
        total_cell = ws[f'A{total_row}']
        total_cell.value = f"Total: {total_packages} paquete{'s' if total_packages != 1 else ''}"
        total_cell.font = Font(bold=True, size=11)
        total_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        total_cell.alignment = Alignment(horizontal='center', vertical='center')
        total_cell.border = border
        
        # Auto-ajustar ancho de columnas
        for col_idx in range(1, len(columns_config) + 1):
            column_letter = get_column_letter(col_idx)
            # Obtener todas las celdas de la columna (saltando filas de encabezado y total)
            data_cells = []
            for row_idx in range(header_row + 1, total_row):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    data_cells.append(cell)
            
            max_length = max(
                (len(str(cell.value)) for cell in data_cells),
                default=10
            )
            # Considerar también el ancho del header
            header_cell = ws.cell(row=header_row, column=col_idx)
            header_width = len(str(header_cell.value)) if header_cell.value else 10
            adjusted_width = min(max(max_length, header_width) + 2, 50)  # Máximo 50 caracteres
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Preparar respuesta HTTP
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"paquetes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @classmethod
    def generate_pdf(cls, queryset: QuerySet["Package"], columns_config: list[str]) -> HttpResponse:
        """
        Genera un archivo PDF con los paquetes especificados
        
        Args:
            queryset: QuerySet de Package
            columns_config: Lista de IDs de campos a incluir en orden
        
        Returns:
            HttpResponse con el archivo PDF
        """
        # Si no hay columnas especificadas, usar las principales
        if not columns_config:
            columns_config = ['guide_number', 'name', 'city', 'status', 'shipment_type_display']
        
        # Limitar columnas para que quepan en PDF (máximo 6 columnas para landscape)
        if len(columns_config) > 6:
            columns_config = columns_config[:6]
        
        # Preparar buffer
        buffer = BytesIO()
        
        # Crear documento en modo landscape para más espacio
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        # Contenedor de elementos
        elements = []
        
        # Contar total de paquetes
        total_packages = queryset.count()
        
        # Agregar título
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        title_style.alignment = 1  # Center alignment
        
        title = Paragraph(
            f"<b>Reporte de Paquetes</b><br/>"
            f"<font size=10>{datetime.now().strftime('%d/%m/%Y %H:%M')}</font>",
            title_style
        )
        elements.append(title)
        elements.append(Paragraph("<br/>", styles['Normal']))
        
        # Agregar información del total
        info_style = styles['Normal']
        info_style.fontSize = 11
        info_style.alignment = 1  # Center alignment
        total_info = Paragraph(
            f"<b>Total de Paquetes: {total_packages}</b>",
            info_style
        )
        elements.append(total_info)
        elements.append(Paragraph("<br/><br/>", styles['Normal']))
        
        # Preparar datos de tabla
        table_data = []
        
        # Headers
        headers = [cls.AVAILABLE_FIELDS.get(col, col) for col in columns_config]
        table_data.append(headers)
        
        # Datos
        for package in queryset:
            row_data = [
                (str(value)[:27] + '...' if len(str(value)) > 30 else value)
                if (value := cls.get_package_field_value(package, field_name)) else ''
                for field_name in columns_config
            ]
            table_data.append(row_data)
        
        # Calcular ancho de columnas dinámicamente
        page_width = landscape(letter)[0] - 60  # Restar márgenes
        col_width = page_width / len(columns_config)
        col_widths = [col_width] * len(columns_config)
        
        # Crear tabla
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        # Estilo de tabla
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Body
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        
        # Agregar resumen al final
        elements.append(Paragraph("<br/>", styles['Normal']))
        summary_style = styles['Normal']
        summary_style.fontSize = 10
        summary_style.alignment = 1  # Center alignment
        summary = Paragraph(
            f"<b>Resumen: {total_packages} paquete{'s' if total_packages != 1 else ''} exportado{'s' if total_packages != 1 else ''}</b>",
            summary_style
        )
        elements.append(summary)
        
        # Construir PDF
        doc.build(elements)
        
        # Preparar respuesta HTTP
        buffer.seek(0)
        filename = f"paquetes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
