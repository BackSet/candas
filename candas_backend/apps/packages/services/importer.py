"""
Servicio para importación de paquetes desde Excel/CSV
"""
from typing import TYPE_CHECKING, Optional
from io import BytesIO
from django.http import HttpResponse
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import openpyxl
import csv
import io

from ..models import Package, PackageImport
from .normalizer import PackageDataNormalizer

if TYPE_CHECKING:
    from django.core.files.uploadedfile import UploadedFile


class PackageImporter:
    """Importador de paquetes desde archivos Excel/CSV"""
    
    @staticmethod
    def preview_file_columns(file: "UploadedFile") -> dict:
        """
        Previsualiza las columnas de un archivo sin importar
        
        Args:
            file: Archivo Excel o CSV
        
        Returns:
            dict: Información de columnas y primeras filas
        """
        try:
            filename = file.name.lower()
            
            if filename.endswith(('.xlsx', '.xls')):
                wb = openpyxl.load_workbook(file, read_only=True)
                ws = wb.active
                
                # Leer headers con comprensión
                headers = [str(cell.value).strip() for cell in ws[1] if cell.value]
                
                # Leer primeras 3 filas de datos (para previsualización)
                preview_rows = [
                    [str(val) if val is not None else '' for val in row]
                    for row in ws.iter_rows(min_row=2, max_row=4, values_only=True)
                    if any(row)
                ]
                
            elif filename.endswith('.csv'):
                file.seek(0)
                content = file.read().decode('utf-8-sig')
                reader = csv.reader(io.StringIO(content))
                headers = [h.strip() for h in next(reader)]
                
                # Leer primeras 3 filas
                preview_rows = []
                for i, row in enumerate(reader):
                    if i >= 3:
                        break
                    preview_rows.append(row)
            
            else:
                return {'error': 'Formato no soportado'}
            
            return {
                'headers': headers,
                'preview_rows': preview_rows,
                'total_columns': len(headers)
            }
            
        except Exception as e:
            return {'error': f'Error al leer archivo: {str(e)}'}
    
    # Mapeo de campos del modelo a nombres en español para la plantilla
    FIELD_LABELS = {
        'guide_number': 'Guía *',
        'name': 'Nombre *',
        'address': 'Dirección *',
        'phone_number': 'Teléfono *',
        'city': 'Ciudad *',
        'province': 'Provincia *',
        'nro_master': 'Número Master',
        'status': 'Estado',
        'notes': 'Notas',
        'hashtags': 'Hashtags',
        'agency_guide_number': 'Guía de Agencia',
        'transport_agency': 'Agencia Transporte',
        'delivery_agency': 'Agencia Reparto',
    }
    
    # Campos obligatorios que siempre deben estar
    REQUIRED_FIELDS = ['guide_number', 'name', 'address', 'phone_number', 'city', 'province']
    
    # Ejemplos de datos para la plantilla
    FIELD_EXAMPLES = {
        'guide_number': 'EC001',
        'name': 'Juan Pérez',
        'address': 'Av. 10 de Agosto y Orellana',
        'phone_number': '0991234567',
        'city': 'QUITO',
        'province': 'PICHINCHA',
        'nro_master': 'MAST-001',
        'status': 'EN_BODEGA',
        'notes': 'Paquete frágil',
        'hashtags': '#urgente #fragil',
        'agency_guide_number': 'AG-2024-001',
        'transport_agency': 'Servientrega',
        'delivery_agency': 'Urbano Express',
    }
    
    @staticmethod
    def generate_template(selected_fields: list[str], field_order: Optional[list[str]] = None) -> HttpResponse:
        """
        Genera una plantilla Excel con los campos seleccionados
        
        Args:
            selected_fields (list): Lista de campos opcionales seleccionados
            field_order (list): Orden personalizado de campos (opcional)
        
        Returns:
            HttpResponse: Archivo Excel descargable
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Importar Paquetes"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        example_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Si hay un orden personalizado, usarlo; si no, usar orden por defecto
        all_fields = (
            [f for f in field_order if f in PackageImporter.REQUIRED_FIELDS or f in selected_fields]
            if field_order
            else PackageImporter.REQUIRED_FIELDS + [f for f in selected_fields if f not in PackageImporter.REQUIRED_FIELDS]
        )
        
        # Fila 1: Headers - usar zip para iteración paralela
        for col, field in enumerate(all_fields, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = PackageImporter.FIELD_LABELS.get(field, field)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Fila 2: Ejemplos - usar zip para iteración paralela
        for col, field in enumerate(all_fields, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = PackageImporter.FIELD_EXAMPLES.get(field, '')
            cell.fill = example_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border
        
        # Ajustar anchos de columna
        for col, field in enumerate(all_fields, 1):
            if field in ['address', 'notes']:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 40
            elif field in ['name', 'transport_agency', 'delivery_agency']:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
            elif field in ['guide_number', 'agency_guide_number', 'nro_master']:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
            else:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Agregar instrucciones en una hoja adicional
        instructions_ws = wb.create_sheet("Instrucciones")
        instructions_ws.column_dimensions['A'].width = 80
        
        instructions = [
            "INSTRUCCIONES PARA IMPORTACIÓN DE PAQUETES",
            "",
            "1. Complete los datos en la hoja 'Importar Paquetes' a partir de la fila 3",
            "2. La fila 2 contiene ejemplos de datos válidos",
            "3. Los campos marcados con * son obligatorios",
            "4. NO modifique los encabezados de la primera fila",
            "",
            "FORMATO DE DATOS:",
            "- Guía: Texto único, sin espacios duplicados",
            "- Teléfono: 10 dígitos, ejemplo: 0991234567",
            "- Ciudad y Provincia: En mayúsculas",
            "- Estado: NO_RECEPTADO, EN_BODEGA, EN_TRANSITO, ENTREGADO, DEVUELTO, RETENIDO",
            "- Hashtags: Separados por espacios, ejemplo: #urgente #fragil",
            "- Agencias: Nombre exacto de la agencia registrada en el sistema",
            "",
            "NORMALIZACIÓN AUTOMÁTICA:",
            "- Los teléfonos se formatearán automáticamente",
            "- Las ciudades y provincias se convertirán a mayúsculas",
            "- Los hashtags se formatearán con # al inicio",
            "- Las guías se limpiarán de espacios y caracteres especiales",
            "",
            "NOTA: Si una fila contiene errores, se omitirá pero el resto se importará correctamente."
        ]
        
        for row, instruction in enumerate(instructions, 1):
            cell = instructions_ws.cell(row=row, column=1)
            cell.value = instruction
            if row == 1:
                cell.font = Font(bold=True, size=14)
            elif instruction.startswith(('1.', '2.', '3.', '4.', 'FORMATO', 'NORMALIZACIÓN', 'NOTA:')):
                cell.font = Font(bold=True)
        
        # Guardar en buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Crear respuesta
        filename = f"plantilla_paquetes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @staticmethod
    def validate_file(file: "UploadedFile", selected_fields: list[str]) -> tuple[bool, Optional[str]]:
        """
        Valida que el archivo tenga el formato correcto
        
        Args:
            file: Archivo cargado
            selected_fields (list): Campos opcionales seleccionados
        
        Returns:
            tuple: (is_valid, error_message)
        """
        try:
            # Determinar si es Excel o CSV
            filename = file.name.lower()
            
            if filename.endswith(('.xlsx', '.xls')):
                wb = openpyxl.load_workbook(file, read_only=True)
                ws = wb.active
                
                # Leer primera fila (headers)
                headers = []
                for cell in ws[1]:
                    if cell.value:
                        headers.append(str(cell.value).strip())
                
            elif filename.endswith('.csv'):
                file.seek(0)
                content = file.read().decode('utf-8-sig')
                reader = csv.reader(io.StringIO(content))
                headers = [h.strip() for h in next(reader)]
                
            else:
                return False, "Formato de archivo no soportado. Use .xlsx, .xls o .csv"
            
            # Validar que tenga headers
            if not headers:
                return False, "El archivo no contiene encabezados"
            
            # Validar que tenga al menos los campos obligatorios
            all_fields = PackageImporter.REQUIRED_FIELDS + selected_fields
            expected_headers = [PackageImporter.FIELD_LABELS.get(f, f) for f in all_fields]
            
            # Verificar campos obligatorios usando all() y comprensión
            missing_fields = [
                PackageImporter.FIELD_LABELS.get(field, field)
                for field in PackageImporter.REQUIRED_FIELDS
                if PackageImporter.FIELD_LABELS.get(field, field) not in headers
            ]
            if missing_fields:
                return False, f"Faltan los campos obligatorios: {', '.join(missing_fields)}"
            
            return True, None
            
        except Exception as e:
            return False, f"Error al validar archivo: {str(e)}"
    
    @staticmethod
    def import_packages(
        file: "UploadedFile",
        selected_fields: list[str],
        import_record_id: int,
        column_mapping: Optional[dict] = None,
        column_order: Optional[dict] = None,
        field_order: Optional[list[str]] = None
    ) -> dict:
        """
        Importa paquetes desde un archivo Excel/CSV
        
        Args:
            file: Archivo a importar
            selected_fields (list): Campos opcionales seleccionados
            import_record_id: ID del registro PackageImport
            column_mapping (dict): Mapeo de índice de columna a campo del modelo
            column_order (dict): Orden de procesamiento de columnas
            field_order (list): Orden personalizado de campos
        
        Returns:
            dict: Resumen de la importación
        """
        import_record = PackageImport.objects.get(id=import_record_id)
        
        try:
            # Validar archivo (si no hay mapeo personalizado)
            if not column_mapping:
                is_valid, error_msg = PackageImporter.validate_file(file, selected_fields)
                if not is_valid:
                    import_record.status = 'ERROR'
                    import_record.error_log = error_msg
                    import_record.save()
                    return {'success': False, 'error': error_msg}
            
            # Leer datos del archivo
            file.seek(0)
            filename = file.name.lower()
            
            if filename.endswith(('.xlsx', '.xls')):
                rows_data = PackageImporter._read_excel(file, selected_fields, column_mapping)
            else:
                rows_data = PackageImporter._read_csv(file, selected_fields, column_mapping)
            
            # Procesar filas
            total_rows = len(rows_data)
            successful = 0
            failed = 0
            error_log = []
            warnings = []
            
            for row_num, row_data in enumerate(rows_data, start=3):  # Start=3 porque row 1=headers, row 2=ejemplos
                try:
                    with transaction.atomic():
                        package, row_warnings = PackageImporter._create_package_from_row(row_data)
                        successful += 1
                        # Agregar advertencias de esta fila usando comprensión
                        warnings.extend(f"Fila {row_num}: {warning}" for warning in row_warnings)
                except Exception as e:
                    failed += 1
                    error_log.append(f"Fila {row_num}: {str(e)}")
            
            # Actualizar registro
            import_record.status = 'COMPLETADO' if failed == 0 else 'COMPLETADO'
            import_record.total_rows = total_rows
            import_record.successful_imports = successful
            import_record.failed_imports = failed
            error_log_str = '\n'.join(error_log) if error_log else ''
            warnings_str = '\n'.join(warnings) if warnings else ''
            # Combinar errores y advertencias en el error_log
            if warnings_str:
                import_record.error_log = f"{error_log_str}\n\nADVERTENCIAS:\n{warnings_str}" if error_log_str else f"ADVERTENCIAS:\n{warnings_str}"
            else:
                import_record.error_log = error_log_str
            import_record.save()
            
            return {
                'success': True,
                'total': total_rows,
                'successful': successful,
                'failed': failed,
                'errors': error_log,
                'warnings': warnings
            }
            
        except Exception as e:
            import_record.status = 'ERROR'
            import_record.error_log = f"Error general: {str(e)}"
            import_record.save()
            return {'success': False, 'error': str(e)}
    
    @staticmethod
    def _read_excel(file: "UploadedFile", selected_fields: list[str], column_mapping: Optional[dict] = None) -> list[dict]:
        """Lee datos de un archivo Excel"""
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        
        # Si hay mapeo personalizado, usarlo
        if column_mapping:
            # column_mapping es un dict como {"0": "guide_number", "1": "name", ...}
            # Convertir índices de string a int
            col_to_field = {int(k): v for k, v in column_mapping.items()}
        else:
            # Leer headers y mapear automáticamente con comprensión
            headers = [str(cell.value).strip() for cell in ws[1] if cell.value]
            
            # Mapear headers a campos del modelo usando comprensión de diccionario
            col_to_field = {
                headers.index(label): field
                for field, label in PackageImporter.FIELD_LABELS.items()
                if label in headers
            }
        
        # Leer datos (saltando fila 1 de headers y fila 2 de ejemplos si es plantilla)
        start_row = 3 if not column_mapping else 2
        rows_data = []
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            # Verificar que la fila no esté completamente vacía
            if not any(row):
                continue
            
            row_dict = {}
            for col_idx, value in enumerate(row):
                if col_idx in col_to_field:
                    field = col_to_field[col_idx]
                    # Convertir valores vacíos a "none"
                    if value is None or (isinstance(value, str) and not value.strip()):
                        row_dict[field] = 'none'
                    else:
                        row_dict[field] = value
            
            rows_data.append(row_dict)
        
        return rows_data
    
    @staticmethod
    def _read_csv(file: "UploadedFile", selected_fields: list[str], column_mapping: Optional[dict] = None) -> list[dict]:
        """Lee datos de un archivo CSV"""
        file.seek(0)
        content = file.read().decode('utf-8-sig')
        
        if column_mapping:
            # Usar mapeo personalizado
            col_to_field = {int(k): v for k, v in column_mapping.items()}
            reader = csv.reader(io.StringIO(content))
            next(reader)  # Saltar header
            
            rows_data = []
            for row in reader:
                row_dict = {}
                for col_idx, value in enumerate(row):
                    if col_idx in col_to_field:
                        field = col_to_field[col_idx]
                        # Convertir valores vacíos a "none"
                        if value is None or (isinstance(value, str) and not value.strip()):
                            row_dict[field] = 'none'
                        else:
                            row_dict[field] = value
                
                if row_dict.get('guide_number'):
                    rows_data.append(row_dict)
        else:
            # Mapeo automático por headers
            reader = csv.DictReader(io.StringIO(content))
            field_to_label = {v: k for k, v in PackageImporter.FIELD_LABELS.items()}
            
            rows_data = []
            for row in reader:
                row_dict = {}
                for label, value in row.items():
                    label = label.strip()
                    if label in field_to_label:
                        field = field_to_label[label]
                        # Convertir valores vacíos a "none"
                        if value is None or (isinstance(value, str) and not value.strip()):
                            row_dict[field] = 'none'
                        else:
                            row_dict[field] = value
                
                # Solo agregar si tiene al menos la guía
                if row_dict.get('guide_number'):
                    rows_data.append(row_dict)
        
        return rows_data
    
    @staticmethod
    def _create_package_from_row(row_data: dict) -> tuple[Package, list[str]]:
        """
        Crea un paquete desde una fila de datos
        
        Args:
            row_data (dict): Datos de la fila
        
        Returns:
            tuple: (Package, list) - Paquete creado y lista de advertencias
        
        Raises:
            Exception: Si hay errores de validación
        """
        normalizer = PackageDataNormalizer
        warnings = []
        
        # Los valores vacíos ya vienen como "none" desde la lectura del archivo
        # Normalizar campos - solo el guide_number es realmente obligatorio
        # Usar operador walrus para simplificar
        if not (guide_number := normalizer.normalize_guide(gn) if (gn := row_data.get('guide_number', 'none')) != 'none' else 'none') or guide_number == 'none':
            raise ValueError("El número de guía es obligatorio")
        
        # Verificar unicidad de guía
        if Package.objects.filter(guide_number=guide_number).exists():
            raise ValueError(f"La guía {guide_number} ya existe en el sistema")
        
        # Campos que antes eran obligatorios ahora se rellenan con "none" si están vacíos
        # Usar operador walrus para simplificar
        name = normalizer.normalize_text(nr) if (nr := row_data.get('name', 'none')) != 'none' else 'none'
        address = normalizer.normalize_address(ar) if (ar := row_data.get('address', 'none')) != 'none' else 'none'
        phone_number = normalizer.normalize_phone(pnr) if (pnr := row_data.get('phone_number', 'none')) != 'none' else 'none'
        
        # Validar teléfono pero no rechazar si tiene menos de 10 dígitos, solo agregar advertencia
        if phone_number != 'none' and not normalizer.validate_phone(phone_number):
            warnings.append(f"Teléfono con formato incompleto: {phone_number} (debe tener 10 dígitos)")
        
        city_raw = row_data.get('city', 'none')
        province_raw = row_data.get('province', 'none')
        city, province = normalizer.normalize_location(
            city_raw if city_raw != 'none' else '',
            province_raw if province_raw != 'none' else ''
        )
        city = 'none' if (city == 'none' or not city) else city
        province = 'none' if (province == 'none' or not province) else province
        
        # Preparar datos del paquete
        # phone_number: si está vacío usar 'none', si tiene valor (aunque incompleto) guardarlo tal cual
        phone_value = 'none' if phone_number == 'none' else phone_number
        
        package_data = {
            'guide_number': guide_number,
            'name': name,
            'address': address,
            'phone_number': phone_value,
            'city': city,
            'province': province,
        }
        
        # Campos opcionales - ya vienen como "none" si estaban vacíos
        # Usar operador walrus para simplificar
        package_data['nro_master'] = (
            normalizer.normalize_text(nmr) if (nmr := row_data.get('nro_master', 'none')) != 'none' else 'none'
        )
        
        package_data['status'] = (
            normalizer.normalize_status(sr) if (sr := row_data.get('status', 'none')) != 'none' else 'NO_RECEPTADO'
        )
        
        package_data['notes'] = (
            normalizer.normalize_text(ntr) if (ntr := row_data.get('notes', 'none')) != 'none' else 'none'
        )
        
        package_data['hashtags'] = (
            normalizer.normalize_hashtags(hr) if (hr := row_data.get('hashtags', 'none')) != 'none' else 'none'
        )
        
        package_data['agency_guide_number'] = (
            normalizer.normalize_guide(agr) if (agr := row_data.get('agency_guide_number', 'none')) != 'none' else 'none'
        )
        
        # Buscar agencias por nombre si se proporcionaron
        if (tar := row_data.get('transport_agency', 'none')) != 'none':
            from apps.catalog.models import TransportAgency
            agency_name = normalizer.normalize_text(tar)
            try:
                package_data['transport_agency'] = TransportAgency.objects.get(name__iexact=agency_name, active=True)
            except TransportAgency.DoesNotExist:
                raise ValueError(f"Agencia de transporte no encontrada: {agency_name}")
        
        if (dar := row_data.get('delivery_agency', 'none')) != 'none':
            from apps.catalog.models import DeliveryAgency
            agency_name = normalizer.normalize_text(dar)
            try:
                package_data['delivery_agency'] = DeliveryAgency.objects.get(name__iexact=agency_name, active=True)
            except DeliveryAgency.DoesNotExist:
                raise ValueError(f"Agencia de reparto no encontrada: {agency_name}")
        
        # Crear paquete
        package = Package.objects.create(**package_data)
        
        return package, warnings
