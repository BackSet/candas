"""
Servicio para generar manifiestos de paquetes individuales.
Formato A4 horizontal, formal, sin colores.
"""
from io import BytesIO
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from datetime import datetime
from apps.shared.services.manifest_template import BaseManifestGenerator


class PackageManifestGenerator:
    """Generador de manifiestos para paquetes individuales."""
    
    @staticmethod
    def generate_pdf(package):
        """
        Genera un PDF con el manifiesto del paquete individual.
        Formato A4 horizontal, formal, sin colores.
        La nota debe estar resaltada.
        """
        doc, buffer = BaseManifestGenerator.create_document()
        elements = []
        
        # Header de la empresa
        elements.append(BaseManifestGenerator.get_company_header())
        elements.append(Spacer(1, 3*mm))
        
        # Título
        title_style = BaseManifestGenerator.get_title_style()
        elements.append(Paragraph("MANIFIESTO DE PAQUETE INDIVIDUAL", title_style))
        elements.append(Spacer(1, 3*mm))
        
        # Información de envío (sin tabla, solo párrafos)
        effective_agency = package.get_shipping_agency()
        effective_guide = package.get_shipping_guide_number()
        effective_destiny = package.get_effective_destiny()
        
        normal_style = BaseManifestGenerator.get_normal_style()
        
        # Crear párrafos separados por saltos de línea
        info_lines = [
            f"Destino: {effective_destiny}",
            f"Agencia de Transporte: {effective_agency.name if effective_agency else 'Sin asignar'}",
            f"Número de Guía de Transporte: {effective_guide or 'Sin guía'}",
            f"Fecha: {BaseManifestGenerator.format_datetime_now()}",
        ]
        
        # Agregar cada línea como un párrafo usando map
        elements.extend(map(lambda line: Paragraph(line, normal_style), info_lines))
        
        elements.append(Spacer(1, 5*mm))
        
        # Datos del paquete
        heading_style = BaseManifestGenerator.get_heading_style()
        elements.append(Paragraph("DATOS DEL PAQUETE", heading_style))
        elements.append(Spacer(1, 3*mm))
        
        # Datos del paquete (sin tabla, solo párrafos)
        package_info_lines = [
            f"Número de Guía del Paquete: {package.guide_number}",
            f"Nombre del Destinatario: {package.name}",
            f"Dirección: {package.address}",
            f"Teléfono: {package.phone_number or '-'}",
        ]
        
        # Agregar número master si existe
        if package.nro_master:
            package_info_lines.insert(1, f"Número Master: {package.nro_master}")
        
        # Agregar cada línea como un párrafo usando map
        elements.extend(map(lambda line: Paragraph(line, normal_style), package_info_lines))
        
        # Nota resaltada
        if package.notes:
            elements.append(Spacer(1, 5*mm))
            
            # Título de nota
            note_heading_style = BaseManifestGenerator.get_heading_style()
            elements.append(Paragraph("NOTA", note_heading_style))
            elements.append(Spacer(1, 2*mm))
            
            # Nota con estilo resaltado (caja con borde)
            note_box = BaseManifestGenerator.create_highlighted_box(package.notes)
            elements.append(note_box)
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        
        return buffer
    
    @staticmethod
    def generate_excel(package):
        """
        Genera un Excel con el manifiesto del paquete individual.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Manifiesto"
        
        # Estilos formales (sin colores)
        header_font = Font(bold=True, size=12)
        normal_font = Font(size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:B1')
        ws['A1'] = 'MANIFIESTO DE PAQUETE INDIVIDUAL'
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Información de envío
        effective_agency = package.get_shipping_agency()
        effective_guide = package.get_shipping_guide_number()
        effective_destiny = package.get_effective_destiny()
        
        row = 3
        ws[f'A{row}'] = 'INFORMACIÓN DE ENVÍO'
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        
        row += 1
        shipping_info = [
            ('Destino:', effective_destiny),
            ('Agencia de Transporte:', effective_agency.name if effective_agency else 'Sin asignar'),
            ('Número de Guía de Transporte:', effective_guide or 'Sin guía'),
            ('Fecha:', datetime.now().strftime('%d/%m/%Y %H:%M')),
        ]
        
        # Usar zip para iteración paralela
        for label, value in shipping_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
        
        # Datos del paquete
        row += 1
        ws[f'A{row}'] = 'DATOS DEL PAQUETE'
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        
        row += 1
        package_info = [
            ('Número de Guía del Paquete:', package.guide_number),
            ('Nombre del Destinatario:', package.name),
            ('Dirección:', package.address),
            ('Teléfono:', package.phone_number or '-'),
        ]
        
        if package.nro_master:
            package_info.insert(1, ('Número Master:', package.nro_master))
        
        # Usar zip para iteración paralela
        for label, value in package_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
        
        # Nota resaltada
        if package.notes:
            row += 1
            ws[f'A{row}'] = 'NOTA'
            ws.merge_cells(f'A{row}:B{row}')
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            
            row += 1
            note_cell = ws[f'A{row}']
            note_cell.value = package.notes
            ws.merge_cells(f'A{row}:B{row}')
            note_cell.font = Font(bold=True, size=11)
            note_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            note_cell.border = border
            note_cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            ws.row_dimensions[row].height = 40
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
