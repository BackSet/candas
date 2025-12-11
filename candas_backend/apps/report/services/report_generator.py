"""
Servicio para generación de reportes en diferentes formatos
"""
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch, mm
import csv


class ReportGenerator:
    """Generador de reportes para diferentes entidades del sistema"""
    
    @staticmethod
    def generate_packages_report(filters=None, format='json'):
        """
        Generar reporte de paquetes con filtros aplicados
        
        Args:
            filters (dict): Filtros a aplicar (date_from, date_to, status, etc.)
            format (str): Formato de salida ('json', 'excel', 'pdf', 'csv')
        
        Returns:
            dict or HttpResponse: Datos del reporte o archivo descargable
        """
        from apps.packages.models import Package
        
        queryset = Package.objects.select_related(
            'transport_agency', 
            'pull', 
            'pull__batch'
        ).all()
        
        # Aplicar filtros
        if filters:
            # Filtro especial: fecha de cambio de estado EN_BODEGA → EN_TRANSITO
            if filters.get('status_change_date'):
                from apps.packages.models import PackageStatusHistory
                from datetime import time, date as date_type
                
                # Convertir fecha string a objeto date si es necesario
                status_change_date = filters['status_change_date']
                if isinstance(status_change_date, str):
                    status_change_date = datetime.fromisoformat(status_change_date.replace('Z', '+00:00')).date()
                elif isinstance(status_change_date, datetime):
                    status_change_date = status_change_date.date()
                
                # Obtener rango de fechas (todo el día)
                start_datetime = datetime.combine(status_change_date, time.min)
                end_datetime = datetime.combine(status_change_date, time.max)
                
                # Buscar cambios de estado EN_BODEGA → EN_TRANSITO en esa fecha
                status_changes = PackageStatusHistory.objects.filter(
                    old_status='EN_BODEGA',
                    new_status='EN_TRANSITO',
                    changed_at__gte=start_datetime,
                    changed_at__lte=end_datetime
                ).values_list('package_id', flat=True)
                
                # Filtrar paquetes que cambiaron de estado en esa fecha
                queryset = queryset.filter(id__in=status_changes)
            
            if filters.get('date_from'):
                queryset = queryset.filter(created_at__gte=filters['date_from'])
            if filters.get('date_to'):
                queryset = queryset.filter(created_at__lte=filters['date_to'])
            if filters.get('status'):
                queryset = queryset.filter(status=filters['status'])
            if filters.get('transport_agency'):
                queryset = queryset.filter(
                    Q(transport_agency_id=filters['transport_agency']) |
                    Q(pull__transport_agency_id=filters['transport_agency']) |
                    Q(pull__batch__transport_agency_id=filters['transport_agency'])
                )
            if filters.get('shipment_type'):
                shipment_type = filters['shipment_type']
                if shipment_type == 'individual':
                    queryset = queryset.filter(pull__isnull=True, transport_agency__isnull=False)
                elif shipment_type == 'saca':
                    queryset = queryset.filter(pull__isnull=False, pull__batch__isnull=True)
                elif shipment_type == 'lote':
                    queryset = queryset.filter(pull__isnull=False, pull__batch__isnull=False)
        
        # Formato de salida
        if format == 'json':
            # Evaluar queryset completamente antes de procesar para obtener conteo correcto
            packages_list = list(queryset.select_related('transport_agency', 'pull', 'pull__batch'))
            data = []
            for package in packages_list:
                agency = package.get_shipping_agency()
                shipment_type = package.get_shipment_type()
                shipment_info = {
                    'type': shipment_type,
                    'type_display': package.get_shipment_type_display(),
                }
                
                # Agregar información de lote si aplica
                if shipment_type == 'lote' and package.pull and package.pull.batch:
                    shipment_info['batch'] = {
                        'id': str(package.pull.batch.id),
                        'destiny': package.pull.batch.destiny,
                        'guide_number': package.pull.batch.guide_number or '',
                    }
                
                # Agregar información de saca si aplica
                if shipment_type in ['saca', 'lote'] and package.pull:
                    shipment_info['pull'] = {
                        'id': str(package.pull.id),
                        'common_destiny': package.pull.get_effective_destiny(),
                        'guide_number': package.pull.get_effective_guide_number() or '',
                    }
                
                data.append({
                    'guide_number': package.guide_number,
                    'name': package.name,
                    'city': package.city,
                    'province': package.province,
                    'status': package.get_status_display(),
                    'shipment_type': package.get_shipment_type_display(),
                    'shipment_info': shipment_info,
                    'agency': agency.name if agency else 'Sin asignar',
                    'created_at': package.created_at.isoformat(),
                })
            # Usar len(data) en lugar de queryset.count() para obtener el conteo correcto
            return {'data': data, 'count': len(data)}
        
        elif format == 'excel':
            # Asegurar que el queryset se evalúe completamente antes de exportar
            # Esto garantiza que todos los filtros se hayan aplicado
            queryset = queryset.select_related('transport_agency', 'pull', 'pull__batch')
            return ReportGenerator.export_packages_to_excel(queryset)
        
        elif format == 'pdf':
            return ReportGenerator.export_packages_to_pdf(queryset)
        
        elif format == 'csv':
            return ReportGenerator.export_packages_to_csv(queryset)
    
    @staticmethod
    def generate_statistics_report(date_from, date_to):
        """
        Generar reporte estadístico general del sistema
        
        Args:
            date_from (date): Fecha desde
            date_to (date): Fecha hasta
        
        Returns:
            dict: Estadísticas generales
        """
        from apps.packages.models import Package
        from apps.logistics.models import Pull, Batch
        from apps.catalog.models import TransportAgency
        
        # Paquetes
        packages = Package.objects.filter(
            created_at__gte=date_from,
            created_at__lte=date_to
        )
        
        packages_by_status = {}
        for status_code, status_name in Package.STATUS_CHOICES:
            count = packages.filter(status=status_code).count()
            packages_by_status[status_code] = {
                'name': status_name,
                'count': count
            }
        
        # Paquetes por tipo de envío
        shipment_types = {
            'individual': packages.filter(pull__isnull=True, transport_agency__isnull=False).count(),
            'saca': packages.filter(pull__isnull=False, pull__batch__isnull=True).count(),
            'lote': packages.filter(pull__isnull=False, pull__batch__isnull=False).count(),
            'sin_envio': packages.filter(pull__isnull=True, transport_agency__isnull=True).count(),
        }
        
        # Sacas y Lotes
        pulls = Pull.objects.filter(created_at__gte=date_from, created_at__lte=date_to)
        batches = Batch.objects.filter(created_at__gte=date_from, created_at__lte=date_to)
        
        # Top agencias
        top_agencies = []
        for agency in TransportAgency.objects.filter(active=True)[:10]:
            agency_packages = packages.filter(
                Q(transport_agency=agency) |
                Q(pull__transport_agency=agency) |
                Q(pull__batch__transport_agency=agency)
            ).count()
            if agency_packages > 0:
                top_agencies.append({
                    'name': agency.name,
                    'packages': agency_packages
                })
        
        top_agencies = sorted(top_agencies, key=lambda x: x['packages'], reverse=True)[:10]
        
        # Top destinos
        top_destinations = packages.values('city').annotate(
            count=Count('id')
        ).order_by('-count')[:10]
        
        return {
            'period': {
                'from': date_from.isoformat(),
                'to': date_to.isoformat(),
            },
            'packages': {
                'total': packages.count(),
                'by_status': packages_by_status,
                'by_shipment_type': shipment_types,
            },
            'pulls': {
                'total': pulls.count(),
            },
            'batches': {
                'total': batches.count(),
            },
            'top_agencies': top_agencies,
            'top_destinations': list(top_destinations),
        }
    
    @staticmethod
    def generate_agencies_performance(date_from, date_to):
        """Reporte de rendimiento de agencias"""
        from apps.catalog.models import TransportAgency
        from apps.packages.models import Package
        
        agencies_data = []
        for agency in TransportAgency.objects.filter(active=True):
            packages = Package.objects.filter(
                Q(transport_agency=agency) |
                Q(pull__transport_agency=agency) |
                Q(pull__batch__transport_agency=agency),
                created_at__gte=date_from,
                created_at__lte=date_to
            )
            
            total = packages.count()
            if total == 0:
                continue
            
            agencies_data.append({
                'name': agency.name,
                'total_packages': total,
                'pulls': agency.pulls.filter(created_at__gte=date_from, created_at__lte=date_to).count(),
                'batches': agency.batches.filter(created_at__gte=date_from, created_at__lte=date_to).count(),
                'by_status': {
                    status_code: packages.filter(status=status_code).count()
                    for status_code, _ in Package.STATUS_CHOICES
                }
            })
        
        agencies_data = sorted(agencies_data, key=lambda x: x['total_packages'], reverse=True)
        
        return {
            'period': {'from': date_from.isoformat(), 'to': date_to.isoformat()},
            'agencies': agencies_data
        }
    
    @staticmethod
    def generate_destinations_report(date_from, date_to):
        """Reporte de distribución por destinos"""
        from apps.packages.models import Package
        
        destinations = Package.objects.filter(
            created_at__gte=date_from,
            created_at__lte=date_to
        ).values('city', 'province').annotate(
            total=Count('id')
        ).order_by('-total')
        
        return {
            'period': {'from': date_from.isoformat(), 'to': date_to.isoformat()},
            'destinations': list(destinations)
        }
    
    @staticmethod
    def export_packages_to_excel(queryset):
        """Exportar paquetes a Excel con formato"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Paquetes"
        
        # Configurar orientación horizontal (landscape)
        # Usar el método set_printer_settings que es más confiable
        ws.set_printer_settings(paper_size=9, orientation='landscape')  # 9 = A4, landscape = horizontal
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Evaluar el queryset completamente antes de contar para obtener el conteo correcto
        # Esto asegura que todos los filtros se hayan aplicado y se eviten múltiples evaluaciones
        packages_list = list(queryset.select_related('transport_agency', 'pull', 'pull__batch'))
        total_packages = len(packages_list)
        
        # Headers
        headers = ['Guía', 'Nombre', 'Ciudad', 'Provincia', 'Estado', 'Tipo Envío', 'Lote/Saca', 'Agencia', 'Fecha']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos
        for row, package in enumerate(packages_list, 2):
            agency = package.get_shipping_agency()
            shipment_type = package.get_shipment_type()
            
            # Determinar información de Lote/Saca
            lote_saca_info = ''
            if shipment_type == 'lote' and package.pull and package.pull.batch:
                lote_saca_info = f"Lote: {package.pull.batch.destiny}"
                if package.pull.get_effective_destiny():
                    lote_saca_info += f" | Saca: {package.pull.get_effective_destiny()}"
            elif shipment_type == 'saca' and package.pull:
                lote_saca_info = f"Saca: {package.pull.get_effective_destiny()}"
            elif shipment_type == 'individual':
                lote_saca_info = 'Paquete Individual'
            else:
                lote_saca_info = 'Sin asignar'
            
            ws.cell(row=row, column=1).value = package.guide_number or ''
            ws.cell(row=row, column=2).value = package.name
            ws.cell(row=row, column=3).value = package.city
            ws.cell(row=row, column=4).value = package.province
            ws.cell(row=row, column=5).value = package.get_status_display()
            ws.cell(row=row, column=6).value = package.get_shipment_type_display()
            ws.cell(row=row, column=7).value = lote_saca_info
            ws.cell(row=row, column=8).value = agency.name if agency else 'Sin asignar'
            ws.cell(row=row, column=9).value = package.created_at.strftime('%Y-%m-%d %H:%M')
            
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = border
        
        # Agregar fila de resumen con el total de paquetes
        summary_row = total_packages + 2
        ws.cell(row=summary_row, column=1).value = f"Total de Paquetes: {total_packages}"
        ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
        ws.merge_cells(f'A{summary_row}:I{summary_row}')
        ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal='right', vertical='center')
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 35
        ws.column_dimensions['H'].width = 25
        ws.column_dimensions['I'].width = 18
        
        # Respuesta
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"reporte_paquetes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    @staticmethod
    def export_packages_to_pdf(queryset):
        """Exportar paquetes a PDF usando formato de manifiestos"""
        from apps.shared.services.manifest_template import BaseManifestGenerator
        
        # Crear documento con formato de manifiesto (A4 horizontal, márgenes estrechos)
        doc, buffer = BaseManifestGenerator.create_document()
        elements = []
        
        # Header de la empresa
        elements.append(BaseManifestGenerator.get_company_header())
        elements.append(Spacer(1, 3*mm))
        
        # Título
        title_style = BaseManifestGenerator.get_title_style()
        elements.append(Paragraph("REPORTE DE PAQUETES", title_style))
        elements.append(Spacer(1, 3*mm))
        
        # Información del reporte (párrafos separados por saltos de línea)
        normal_style = BaseManifestGenerator.get_normal_style()
        
        # Evaluar queryset completamente para obtener conteo correcto
        packages_list = list(queryset.select_related('transport_agency', 'pull', 'pull__batch'))
        total_packages = len(packages_list)
        
        info_lines = [
            f"Fecha de Generación: {BaseManifestGenerator.format_datetime_now()}",
            f"Total de Paquetes: {total_packages}",
        ]
        
        # Agregar cada línea como un párrafo
        for line in info_lines:
            elements.append(Paragraph(line, normal_style))
        
        elements.append(Spacer(1, 5*mm))
        
        # Lista de paquetes
        heading_style = BaseManifestGenerator.get_heading_style()
        elements.append(Paragraph("LISTA DE PAQUETES", heading_style))
        elements.append(Spacer(1, 3*mm))
        
        if packages_list:
            # Headers de la tabla
            headers = ['#', 'Número de Guía', 'Nombre', 'Ciudad', 'Provincia', 'Estado', 'Tipo Envío', 'Lote/Saca', 'Agencia']
            
            # Preparar datos
            pkg_data = []
            for idx, package in enumerate(packages_list, 1):
                agency = package.get_shipping_agency()
                shipment_type = package.get_shipment_type()
                
                # Determinar información de Lote/Saca
                lote_saca_info = ''
                if shipment_type == 'lote' and package.pull and package.pull.batch:
                    lote_saca_info = f"Lote: {package.pull.batch.destiny}"
                    if package.pull.get_effective_destiny():
                        lote_saca_info += f" | Saca: {package.pull.get_effective_destiny()}"
                elif shipment_type == 'saca' and package.pull:
                    lote_saca_info = f"Saca: {package.pull.get_effective_destiny()}"
                elif shipment_type == 'individual':
                    lote_saca_info = 'Paquete Individual'
                else:
                    lote_saca_info = 'Sin asignar'
                
                pkg_data.append([
                    str(idx),
                    package.guide_number or '-',
                    package.name or '-',
                    package.city or '-',
                    package.province or '-',
                    package.get_status_display(),
                    package.get_shipment_type_display(),
                    lote_saca_info,
                    agency.name if agency else 'Sin asignar',
                ])
            
            # Crear tabla usando el formato de manifiestos
            # Anchos de columnas ajustados para A4 horizontal con márgenes de 10mm
            available_width = BaseManifestGenerator.PAGE_SIZE[0] - (2 * BaseManifestGenerator.MARGIN)  # 297mm - 20mm = 277mm
            col_widths = [
                10*mm,      # #
                35*mm,      # Número de Guía
                40*mm,      # Nombre
                30*mm,      # Ciudad
                30*mm,      # Provincia
                25*mm,      # Estado
                30*mm,      # Tipo Envío
                40*mm,      # Lote/Saca
                37*mm       # Agencia
            ]
            
            # Ajustar para que sumen exactamente el ancho disponible
            total_width = sum(col_widths)
            if total_width != available_width:
                ratio = available_width / total_width
                col_widths = [w * ratio for w in col_widths]
            
            packages_table = BaseManifestGenerator.create_data_table(
                headers, pkg_data, col_widths
            )
            elements.append(packages_table)
        else:
            elements.append(Paragraph("No hay paquetes para mostrar", normal_style))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"reporte_paquetes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    @staticmethod
    def export_packages_to_csv(queryset):
        """Exportar paquetes a CSV"""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f"reporte_paquetes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # BOM para Excel
        response.write('\ufeff')
        
        writer = csv.writer(response)
        writer.writerow(['Guía', 'Nombre', 'Ciudad', 'Provincia', 'Estado', 'Tipo Envío', 'Lote/Saca', 'Agencia', 'Fecha'])
        
        for package in queryset:
            agency = package.get_shipping_agency()
            shipment_type = package.get_shipment_type()
            
            # Determinar información de Lote/Saca
            lote_saca_info = ''
            if shipment_type == 'lote' and package.pull and package.pull.batch:
                lote_saca_info = f"Lote: {package.pull.batch.destiny}"
                if package.pull.get_effective_destiny():
                    lote_saca_info += f" | Saca: {package.pull.get_effective_destiny()}"
            elif shipment_type == 'saca' and package.pull:
                lote_saca_info = f"Saca: {package.pull.get_effective_destiny()}"
            elif shipment_type == 'individual':
                lote_saca_info = 'Paquete Individual'
            else:
                lote_saca_info = 'Sin asignar'
            
            writer.writerow([
                package.guide_number or '',
                package.name,
                package.city,
                package.province,
                package.get_status_display(),
                package.get_shipment_type_display(),
                lote_saca_info,
                agency.name if agency else 'Sin asignar',
                package.created_at.strftime('%Y-%m-%d %H:%M'),
            ])
        
        return response
