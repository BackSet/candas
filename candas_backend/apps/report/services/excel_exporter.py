from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Exportador de informes a formato Excel usando openpyxl.
    """
    
    def __init__(self, report):
        """
        Inicializa el exportador con un informe.
        
        Args:
            report: Instancia de Report
        """
        self.report = report
        self.workbook = Workbook()
        # Eliminar la hoja por defecto
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
    
    def generate(self):
        """
        Genera el archivo Excel y retorna un objeto BytesIO.
        
        Returns:
            BytesIO: Buffer con el contenido del Excel
        """
        logger.info(f"Generando Excel para informe {self.report.id}")
        
        # Crear hojas
        self._create_summary_sheet()
        self._create_agency_sheet()
        self._create_destination_sheet()
        self._create_status_sheet()
        self._create_packages_sheet()
        self._create_pulls_sheet()
        self._create_batches_sheet()
        
        # Guardar en buffer
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        
        logger.info(f"Excel generado exitosamente para informe {self.report.id}")
        return buffer
    
    def _create_summary_sheet(self):
        """Crea la hoja de resumen general."""
        ws = self.workbook.create_sheet("Resumen General")
        
        # Título
        ws['A1'] = self._get_report_title()
        ws['A1'].font = Font(size=16, bold=True, color='1a365d')
        ws.merge_cells('A1:B1')
        
        # Información del periodo
        ws['A2'] = self._get_period_info()
        ws.merge_cells('A2:B2')
        
        # Cabecera de resumen
        ws['A4'] = 'Métrica'
        ws['B4'] = 'Cantidad'
        self._apply_header_style(ws, 'A4:B4')
        
        # Datos de resumen
        data = [
            ['Total de Paquetes', self.report.total_packages],
            ['Paquetes Individuales', self.report.total_individual_packages],
            ['Paquetes en Sacas', self.report.total_packages - self.report.total_individual_packages],
            ['Total de Sacas', self.report.total_sacas],
            ['Total de Lotes', self.report.total_lotes],
        ]
        
        for idx, (metric, value) in enumerate(data, start=5):
            ws[f'A{idx}'] = metric
            ws[f'B{idx}'] = value
            self._apply_data_style(ws, f'A{idx}:B{idx}')
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
    
    def _create_agency_sheet(self):
        """Crea la hoja de desglose por agencia."""
        ws = self.workbook.create_sheet("Por Agencia")
        
        # Título
        ws['A1'] = 'Desglose por Agencia de Transporte'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        # Cabeceras
        headers = ['Agencia', 'Paquetes', 'Sacas', 'Lotes']
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2c5282', end_color='2c5282', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        if self.report.json_data and 'by_agency' in self.report.json_data:
            row = 4
            by_agency = self.report.json_data['by_agency']
            for agency_name, metrics in sorted(by_agency.items()):
                ws.cell(row=row, column=1, value=agency_name)
                ws.cell(row=row, column=2, value=metrics['packages_count'])
                ws.cell(row=row, column=3, value=metrics['sacas_count'])
                ws.cell(row=row, column=4, value=metrics['lotes_count'])
                row += 1
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 35
        for col in ['B', 'C', 'D']:
            ws.column_dimensions[col].width = 12
    
    def _create_destination_sheet(self):
        """Crea la hoja de desglose por destino."""
        ws = self.workbook.create_sheet("Por Destino")
        
        # Título
        ws['A1'] = 'Desglose por Destino'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        # Cabeceras
        headers = ['Destino', 'Paquetes', 'Sacas', 'Lotes']
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2c5282', end_color='2c5282', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        if self.report.json_data and 'by_destination' in self.report.json_data:
            row = 4
            by_destination = self.report.json_data['by_destination']
            # Ordenar por cantidad de paquetes
            sorted_destinations = sorted(
                by_destination.items(),
                key=lambda x: x[1]['packages_count'],
                reverse=True
            )
            
            for destination, metrics in sorted_destinations:
                ws.cell(row=row, column=1, value=destination)
                ws.cell(row=row, column=2, value=metrics['packages_count'])
                ws.cell(row=row, column=3, value=metrics['sacas_count'])
                ws.cell(row=row, column=4, value=metrics['lotes_count'])
                row += 1
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 40
        for col in ['B', 'C', 'D']:
            ws.column_dimensions[col].width = 12
    
    def _create_status_sheet(self):
        """Crea la hoja de desglose por estado."""
        ws = self.workbook.create_sheet("Por Estado")
        
        # Título
        ws['A1'] = 'Desglose por Estado de Paquetes'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        # Cabeceras
        headers = ['Estado', 'Total', 'Individuales', 'En Sacas']
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2c5282', end_color='2c5282', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        if self.report.json_data and 'by_status' in self.report.json_data:
            row = 4
            by_status = self.report.json_data['by_status']
            for status, metrics in sorted(by_status.items()):
                ws.cell(row=row, column=1, value=status)
                ws.cell(row=row, column=2, value=metrics['count'])
                ws.cell(row=row, column=3, value=metrics['individual_count'])
                ws.cell(row=row, column=4, value=metrics['in_pull_count'])
                row += 1
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D']:
            ws.column_dimensions[col].width = 12
    
    def _create_packages_sheet(self):
        """Crea la hoja con el detalle de paquetes."""
        ws = self.workbook.create_sheet("Detalle Paquetes")
        
        # Título
        ws['A1'] = 'Detalle de Paquetes'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:J1')
        
        # Cabeceras
        headers = [
            'Nro. Guía', 'Guía de Envío', 'Nombre', 'Ciudad', 'Provincia',
            'Estado', 'Tipo Envío', 'Agencia', 'Fecha Creación'
        ]
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2c5282', end_color='2c5282', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        if self.report.json_data and 'details' in self.report.json_data:
            packages = self.report.json_data['details'].get('packages', [])
            row = 4
            
            for package in packages:
                ws.cell(row=row, column=1, value=package['guide_number'])
                ws.cell(row=row, column=2, value=package['shipping_guide_number'])
                ws.cell(row=row, column=3, value=package['name'])
                ws.cell(row=row, column=4, value=package['city'])
                ws.cell(row=row, column=5, value=package['province'])
                ws.cell(row=row, column=6, value=package['status_display'])
                ws.cell(row=row, column=7, value='Individual' if package['is_individual'] else 'En Saca')
                ws.cell(row=row, column=8, value=package['transport_agency'] or 'Sin Agencia')
                ws.cell(row=row, column=9, value=package['created_at'][:10])
                row += 1
        
        # Ajustar ancho de columnas
        for col_idx in range(1, 10):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        ws.column_dimensions['C'].width = 25  # Nombre
    
    def _create_pulls_sheet(self):
        """Crea la hoja con el detalle de sacas."""
        ws = self.workbook.create_sheet("Detalle Sacas")
        
        # Título
        ws['A1'] = 'Detalle de Sacas'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        
        # Cabeceras
        headers = ['Nro. Guía', 'Destino', 'Tamaño', 'Agencia', 'Cant. Paquetes', 'Fecha Creación']
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2c5282', end_color='2c5282', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        if self.report.json_data and 'details' in self.report.json_data:
            pulls = self.report.json_data['details'].get('pulls', [])
            row = 4
            
            for pull in pulls:
                ws.cell(row=row, column=1, value=pull['guide_number'])
                ws.cell(row=row, column=2, value=pull['common_destiny'])
                ws.cell(row=row, column=3, value=pull['size_display'])
                ws.cell(row=row, column=4, value=pull['transport_agency'] or 'Sin Agencia')
                ws.cell(row=row, column=5, value=pull['packages_count'])
                ws.cell(row=row, column=6, value=pull['created_at'][:10])
                row += 1
        
        # Ajustar ancho de columnas
        for col_idx in range(1, 7):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
    
    def _create_batches_sheet(self):
        """Crea la hoja con el detalle de lotes."""
        ws = self.workbook.create_sheet("Detalle Lotes")
        
        # Título
        ws['A1'] = 'Detalle de Lotes'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        
        # Cabeceras
        headers = ['Nro. Guía', 'Destino', 'Agencia', 'Cant. Sacas', 'Cant. Paquetes', 'Fecha Creación']
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2c5282', end_color='2c5282', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        if self.report.json_data and 'details' in self.report.json_data:
            batches = self.report.json_data['details'].get('batches', [])
            row = 4
            
            for batch in batches:
                ws.cell(row=row, column=1, value=batch['guide_number'])
                ws.cell(row=row, column=2, value=batch['destiny'])
                ws.cell(row=row, column=3, value=batch['transport_agency'] or 'Sin Agencia')
                ws.cell(row=row, column=4, value=batch['pulls_count'])
                ws.cell(row=row, column=5, value=batch['packages_count'])
                ws.cell(row=row, column=6, value=batch['created_at'][:10])
                row += 1
        
        # Ajustar ancho de columnas
        for col_idx in range(1, 7):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
    
    def _get_report_title(self):
        """Retorna el título del informe."""
        if self.report.report_type == 'DAILY':
            return f"Informe Diario de Envíos - {self.report.report_date.strftime('%d/%m/%Y')}"
        else:
            return f"Informe Mensual de Envíos - {self.report.report_date.strftime('%m/%Y')}"
    
    def _get_period_info(self):
        """Retorna información del periodo del informe."""
        generated_at = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        generated_by = self.report.generated_by.username if self.report.generated_by else 'Sistema Automático'
        return f"Generado el {generated_at} por {generated_by}"
    
    def _apply_header_style(self, ws, cell_range):
        """Aplica estilo de cabecera a un rango de celdas."""
        for row in ws[cell_range]:
            for cell in row:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='2c5282', end_color='2c5282', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _apply_data_style(self, ws, cell_range):
        """Aplica estilo de datos a un rango de celdas."""
        for row in ws[cell_range]:
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')
