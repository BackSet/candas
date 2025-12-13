"""
ViewSets para el módulo de Logistics
"""
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import get_object_or_404
from django.http import FileResponse, HttpResponse
from datetime import datetime
from rest_framework.parsers import MultiPartParser, FormParser
from ..models import Pull, Batch, Dispatch
from ..services import PullService, PDFService, QRService, BatchManifestGenerator, BatchLabelsGenerator
from apps.shared.services.universal_manifest_generator import UniversalManifestGenerator
from apps.shared.services.manifest_registry import ManifestRegistry
from .serializers import (
    PullListSerializer,
    PullDetailSerializer,
    PullCreateSerializer,
    PullBulkCreateSerializer,
    BatchSerializer,
    BatchListSerializer,
    BatchDetailSerializer,
    BatchCreateSerializer,
    BatchWithPullsCreateSerializer,
    BatchAutoDistributeSerializer,
    DispatchSerializer,
    DispatchDetailSerializer,
)


class PullViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar Pulls
    """
    queryset = Pull.objects.select_related(
        'transport_agency',
        'batch'
    ).prefetch_related('packages')
    permission_classes = [IsAuthenticated]
    search_fields = ['common_destiny', 'guide_number']
    ordering_fields = ['created_at', 'updated_at', 'common_destiny']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        if self.action == 'list':
            return PullListSerializer
        elif self.action == 'create':
            return PullCreateSerializer
        return PullDetailSerializer
    
    def get_queryset(self):
        """Filtrar queryset"""
        queryset = Pull.objects.all()
        
        # Filtro por batch
        batch_id = self.request.query_params.get('batch', None)
        if batch_id:
            queryset = queryset.filter(batch_id=batch_id)
        
        # Filtro para sacas sin lote (individuales)
        available = self.request.query_params.get('available', None)
        if available == 'true' or available is True:
            queryset = queryset.filter(batch__isnull=True)
        
        # Filtro por agencia
        agency_id = self.request.query_params.get('agency', None)
        if agency_id:
            queryset = queryset.filter(transport_agency_id=agency_id)
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def add_packages(self, request, pk=None):
        """
        Agregar paquetes a un Pull
        POST /api/v1/pulls/{id}/add_packages/
        Body: {"package_ids": ["uuid1", "uuid2", ...]}
        """
        pull = self.get_object()
        package_ids = request.data.get('package_ids', [])
        
        if not package_ids:
            return Response(
                {'error': 'package_ids es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from apps.packages.models import Package
            packages = Package.objects.filter(id__in=package_ids, pull__isnull=True)
            
            if packages.count() != len(package_ids):
                return Response(
                    {'error': 'Algunos paquetes no están disponibles'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Asignar paquetes
            packages.update(pull=pull)
            
            pull.refresh_from_db()
            serializer = PullDetailSerializer(pull, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def remove_packages(self, request, pk=None):
        """
        Remover paquetes de un Pull
        POST /api/v1/pulls/{id}/remove_packages/
        Body: {"package_ids": ["uuid1", "uuid2", ...]}
        """
        pull = self.get_object()
        package_ids = request.data.get('package_ids', [])
        
        if not package_ids:
            return Response(
                {'error': 'package_ids es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        from apps.packages.models import Package
        packages = Package.objects.filter(id__in=package_ids, pull=pull)
        packages.update(pull=None)
        
        pull.refresh_from_db()
        serializer = PullDetailSerializer(pull, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'], url_path='change-packages-status')
    def change_packages_status(self, request, pk=None):
        """
        Cambiar el estado de todos los paquetes de una saca.
        POST /api/v1/pulls/{id}/change-packages-status/
        Body: {"status": "EN_TRANSITO"}
        """
        pull = self.get_object()
        new_status = request.data.get('status', '').strip()
        
        if not new_status:
            return Response(
                {'error': 'El campo status es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar que el estado sea válido
        from apps.packages.models import Package
        valid_statuses = [choice[0] for choice in Package.STATUS_CHOICES]
        
        if new_status not in valid_statuses:
            return Response(
                {'error': f'Estado inválido. Estados válidos: {", ".join(valid_statuses)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Obtener todos los paquetes de la saca
            packages = Package.objects.filter(pull=pull)
            total_packages = packages.count()
            
            if total_packages == 0:
                return Response(
                    {'error': 'La saca no tiene paquetes asociados'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Cambiar el estado de todos los paquetes
            # Usamos un loop para que se disparen los signals/save de cada paquete
            updated_count = 0
            for package in packages:
                if package.status != new_status:
                    package.status = new_status
                    package.save()
                    updated_count += 1
            
            return Response({
                'message': f'Estado actualizado a {new_status}',
                'total_packages': total_packages,
                'updated_count': updated_count,
                'already_in_status': total_packages - updated_count
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': f'Error al actualizar estados: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def statistics(self, request, pk=None):
        """
        Obtener estadísticas de un Pull
        GET /api/v1/pulls/{id}/statistics/
        """
        pull = self.get_object()
        stats = PullService.get_pull_statistics(pull)
        return Response(stats)
    
    @action(detail=True, methods=['post'])
    def generate_barcode(self, request, pk=None):
        """
        Generar código de barras para un Pull
        POST /api/v1/pulls/{id}/generate_barcode/
        """
        pull = self.get_object()
        try:
            PullService.create_barcode(pull)
            pull.refresh_from_db()
            serializer = PullDetailSerializer(pull, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """
        Crear múltiples sacas con el mismo destino
        POST /api/v1/pulls/bulk_create/
        Body: {
            "common_destiny": "CIUDAD - DESTINO",
            "transport_agency": "uuid",
            "guide_number": "GUIA-001",
            "pulls": [
                {"size": "PEQUENO", "package_ids": ["uuid1", "uuid2"]},
                {"size": "MEDIANO", "package_ids": ["uuid3", "uuid4"]}
            ]
        }
        """
        serializer = PullBulkCreateSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(
                serializer.errors,
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            created_pulls = serializer.save()
            
            # Serializar las sacas creadas
            pulls_data = PullListSerializer(
                created_pulls,
                many=True,
                context={'request': request}
            ).data
            
            return Response({
                'message': f'{len(created_pulls)} saca(s) creada(s) exitosamente',
                'pulls': pulls_data,
                'count': len(created_pulls)
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['get'])
    def generate_manifest(self, request, pk=None):
        """
        Generar PDF con manifiesto de paquetes (sistema automático)
        GET /api/v1/pulls/{id}/generate_manifest/
        """
        pull = self.get_object()
        
        try:
            # Obtener adapter automáticamente
            adapter = ManifestRegistry.get_adapter_for_entity(pull)
            
            # Generar PDF usando sistema universal
            pdf_buffer = UniversalManifestGenerator.generate_pdf(pull, adapter)
            
            # Nombre del archivo
            filename = f"manifiesto_saca_{pull.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            # Retornar PDF
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Error al generar manifiesto: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def generate_manifest_excel(self, request, pk=None):
        """
        Generar Excel con manifiesto de la saca (sistema automático)
        GET /api/v1/pulls/{id}/generate_manifest_excel/
        """
        pull = self.get_object()
        
        try:
            # Obtener adapter automáticamente
            adapter = ManifestRegistry.get_adapter_for_entity(pull)
            
            # Generar Excel usando sistema universal
            excel_buffer = UniversalManifestGenerator.generate_excel(pull, adapter)
            
            # Nombre del archivo
            filename = f"manifiesto_saca_{pull.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Retornar Excel
            response = HttpResponse(
                excel_buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Error al generar manifiesto Excel: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def generate_label(self, request, pk=None):
        """
        Generar PDF con etiqueta de saca (1/3 de hoja A4)
        GET /api/v1/pulls/{id}/generate_label/
        """
        pull = self.get_object()
        
        try:
            # Generar PDF
            pdf_buffer = PDFService.generate_pull_label(pull)
            
            # Nombre del archivo
            filename = f"etiqueta_saca_{pull.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            # Retornar PDF
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Error al generar etiqueta: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def qr_code(self, request, pk=None):
        """
        Generar código QR en base64
        GET /api/v1/pulls/{id}/qr_code/
        """
        pull = self.get_object()
        
        try:
            # Generar QR en base64
            qr_base64 = QRService.generate_pull_qr_base64(pull)
            
            return Response({
                'qr_code': qr_base64,
                'pull_id': str(pull.id),
                'common_destiny': pull.common_destiny
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': f'Error al generar código QR: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def generate_notification_message(self, request, pk=None):
        """
        Generar mensaje de notificación para avisar al usuario que su paquete ha sido enviado.
        GET /api/v1/pulls/{id}/generate_notification_message/
        """
        pull = self.get_object()
        
        try:
            from apps.packages.models import Package
            
            # Obtener información efectiva de la saca
            transport_agency = pull.get_effective_agency()
            guide_number = pull.get_effective_guide_number()
            total_packages = Package.objects.filter(pull=pull).count()
            
            if not transport_agency:
                return Response(
                    {'error': 'La saca no tiene agencia de transporte asignada'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if not guide_number:
                return Response(
                    {'error': 'La saca no tiene número de guía asignado'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Determinar URL de seguimiento según la agencia
            tracking_url = None
            agency_name_lower = transport_agency.name.lower()
            if 'yobel' in agency_name_lower:
                tracking_url = 'https://www.yobelscm.biz/ecuador/tracking-clientes-yobel-express-2/'
            
            # Generar mensaje
            message_parts = [
                f"Su paquete ha sido enviado a través de la agencia {transport_agency.name}.",
                f"Número de guía: {guide_number}.",
            ]
            
            if total_packages > 0:
                message_parts.append(f"Se han enviado {total_packages} paquete(s).")
            
            if tracking_url:
                message_parts.append(f"Puede realizar el seguimiento en: {tracking_url}")
            
            message = " ".join(message_parts)
            
            return Response({
                'message': message,
                'agency_name': transport_agency.name,
                'guide_number': guide_number,
                'total_packages': total_packages,
                'tracking_url': tracking_url,
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': f'Error al generar mensaje de notificación: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class BatchViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar Batches
    """
    queryset = Batch.objects.select_related('transport_agency').prefetch_related('pulls')
    permission_classes = [IsAuthenticated]
    search_fields = ['destiny', 'guide_number']
    ordering_fields = ['created_at', 'destiny']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        """Usar serializer diferente según la acción"""
        if self.action == 'list':
            return BatchListSerializer
        elif self.action == 'retrieve':
            return BatchDetailSerializer
        elif self.action == 'create':
            return BatchCreateSerializer
        return BatchSerializer
    
    def get_queryset(self):
        """Filtrar queryset según parámetros"""
        queryset = Batch.objects.select_related('transport_agency').prefetch_related('pulls')
        
        # Filtro por agencia
        agency_id = self.request.query_params.get('transport_agency', None)
        if agency_id:
            queryset = queryset.filter(transport_agency_id=agency_id)
        
        # Filtro por destino
        destiny = self.request.query_params.get('destiny', None)
        if destiny:
            queryset = queryset.filter(destiny__icontains=destiny)
        
        # Búsqueda general
        search = self.request.query_params.get('search', None)
        if search:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(destiny__icontains=search) |
                Q(guide_number__icontains=search)
            )
        
        return queryset
    
    @action(detail=False, methods=['post'])
    def create_with_pulls(self, request):
        """
        Crear lote y múltiples sacas en una operación
        POST /api/v1/batches/create_with_pulls/
        Body: {
            "destiny": "CIUDAD",
            "transport_agency": "uuid",
            "guide_number": "LOTE-001",
            "pulls": [
                {"size": "PEQUENO", "package_ids": [...]},
                {"size": "MEDIANO", "package_ids": [...]}
            ]
        }
        """
        serializer = BatchWithPullsCreateSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(
                serializer.errors,
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            batch = serializer.save()
            
            # Serializar el batch creado
            batch_data = BatchSerializer(batch, context={'request': request}).data
            
            # Agregar información de las sacas creadas
            pulls_data = PullListSerializer(
                batch.created_pulls,
                many=True,
                context={'request': request}
            ).data
            
            return Response({
                'message': f'Lote creado con {len(batch.created_pulls)} saca(s)',
                'batch': batch_data,
                'pulls': pulls_data,
                'pulls_count': len(batch.created_pulls)
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['post'])
    def auto_distribute(self, request):
        """
        Crear lote y distribuir paquetes automáticamente
        POST /api/v1/batches/auto_distribute/
        Body: {
            "destiny": "CIUDAD",
            "transport_agency": "uuid",
            "guide_number": "LOTE-001",
            "package_ids": ["uuid1", "uuid2", ...],
            "pulls_config": [
                {"size": "PEQUENO", "max_packages": 10},
                {"size": "MEDIANO", "max_packages": 20}
            ]
        }
        """
        serializer = BatchAutoDistributeSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(
                serializer.errors,
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            batch = serializer.save()
            
            # Serializar el batch creado
            batch_data = BatchSerializer(batch, context={'request': request}).data
            
            # Agregar información de las sacas creadas
            pulls_data = PullListSerializer(
                batch.created_pulls,
                many=True,
                context={'request': request}
            ).data
            
            return Response({
                'message': f'Lote creado con {len(batch.created_pulls)} saca(s)',
                'batch': batch_data,
                'pulls': pulls_data,
                'pulls_count': len(batch.created_pulls),
                'distributed_packages': batch.distributed_packages,
                'total_packages': len(request.data.get('package_ids', []))
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['get'])
    def packages_summary(self, request, pk=None):
        """
        Obtener resumen de paquetes del lote agrupados por estado
        GET /api/v1/batches/{id}/packages_summary/
        """
        batch = self.get_object()
        
        from apps.packages.models import Package
        packages = Package.objects.filter(pull__batch=batch)
        
        # Resumen por estado
        status_summary = {}
        for status_code, status_name in Package.STATUS_CHOICES:
            count = packages.filter(status=status_code).count()
            status_summary[status_code] = {
                'name': status_name,
                'count': count
            }
        
        return Response({
            'batch_id': str(batch.id),
            'total_packages': packages.count(),
            'total_pulls': batch.pulls.count(),
            'status_summary': status_summary
        })
    
    @action(detail=True, methods=['post'], url_path='change-packages-status')
    def change_packages_status(self, request, pk=None):
        """
        Cambiar el estado de todos los paquetes de un lote (todas las sacas).
        POST /api/v1/batches/{id}/change-packages-status/
        Body: {"status": "EN_TRANSITO"}
        """
        batch = self.get_object()
        new_status = request.data.get('status', '').strip()
        
        if not new_status:
            return Response(
                {'error': 'El campo status es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar que el estado sea válido
        from apps.packages.models import Package
        valid_statuses = [choice[0] for choice in Package.STATUS_CHOICES]
        
        if new_status not in valid_statuses:
            return Response(
                {'error': f'Estado inválido. Estados válidos: {", ".join(valid_statuses)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Obtener todos los paquetes de todas las sacas del lote
            packages = Package.objects.filter(pull__batch=batch)
            total_packages = packages.count()
            
            if total_packages == 0:
                return Response(
                    {'error': 'El lote no tiene paquetes asociados'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Cambiar el estado de todos los paquetes
            # Usamos un loop para que se disparen los signals/save de cada paquete
            updated_count = 0
            for package in packages:
                if package.status != new_status:
                    package.status = new_status
                    package.save()
                    updated_count += 1
            
            return Response({
                'message': f'Estado actualizado a {new_status}',
                'total_packages': total_packages,
                'updated_count': updated_count,
                'already_in_status': total_packages - updated_count,
                'pulls_affected': batch.pulls.count()
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': f'Error al actualizar estados: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def add_pull(self, request, pk=None):
        """
        Agregar saca existente al lote
        POST /api/v1/batches/{id}/add_pull/
        Body: {"pull_id": "uuid"}
        """
        batch = self.get_object()
        pull_id = request.data.get('pull_id')
        
        if not pull_id:
            return Response(
                {'error': 'pull_id es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            pull = get_object_or_404(Pull, pk=pull_id)
            
            # Validar que la saca no esté en otro lote
            if pull.batch and pull.batch.id != batch.id:
                return Response(
                    {'error': f'La saca ya pertenece al lote {pull.batch.id}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validar que el destino coincida
            if pull.common_destiny != batch.destiny:
                return Response(
                    {'error': f'El destino de la saca ({pull.common_destiny}) no coincide con el destino del lote ({batch.destiny})'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Asignar al lote
            pull.batch = batch
            pull.save()
            
            batch.refresh_from_db()
            serializer = BatchDetailSerializer(batch, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)
            
        except Pull.DoesNotExist:
            return Response(
                {'error': 'Saca no encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def remove_pull(self, request, pk=None):
        """
        Quitar saca del lote
        POST /api/v1/batches/{id}/remove_pull/
        Body: {"pull_id": "uuid"}
        """
        batch = self.get_object()
        pull_id = request.data.get('pull_id')
        
        if not pull_id:
            return Response(
                {'error': 'pull_id es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            pull = get_object_or_404(Pull, pk=pull_id, batch=batch)
            pull.batch = None
            pull.save()
            
            batch.refresh_from_db()
            serializer = BatchDetailSerializer(batch, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)
            
        except Pull.DoesNotExist:
            return Response(
                {'error': 'Saca no encontrada en este lote'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['post'])
    def export(self, request):
        """
        Exportar lista de lotes a Excel
        POST /api/v1/batches/export/
        Body: {"format": "excel"}
        """
        export_format = request.data.get('format', 'excel')
        queryset = self.get_queryset()
        
        from openpyxl import Workbook
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Lotes"
        
        # Headers
        headers = ['ID', 'Destino', 'Agencia', 'Nº Guía', 'Sacas', 'Paquetes', 'Fecha Creación']
        ws.append(headers)
        
        # Datos
        for batch in queryset:
            ws.append([
                str(batch.id)[:8],
                batch.destiny,
                batch.transport_agency.name if batch.transport_agency else 'N/A',
                batch.guide_number or '',
                batch.pulls.count(),
                batch.get_total_packages(),
                batch.created_at.strftime('%Y-%m-%d %H:%M'),
            ])
        
        # Auto-ajustar columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Respuesta
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"lotes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    @action(detail=True, methods=['get'])
    def generate_manifest_pdf(self, request, pk=None):
        """
        Generar PDF con manifiesto del lote (sistema automático)
        GET /api/v1/batches/{id}/generate_manifest_pdf/
        """
        batch = self.get_object()
        
        try:
            # Obtener adapter automáticamente
            adapter = ManifestRegistry.get_adapter_for_entity(batch)
            
            # Generar PDF usando sistema universal
            pdf_buffer = UniversalManifestGenerator.generate_pdf(batch, adapter)
            
            # Nombre del archivo
            filename = f"manifiesto_lote_{str(batch.id)[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            # Retornar PDF
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Error al generar manifiesto PDF: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def generate_manifest_excel(self, request, pk=None):
        """
        Generar Excel con manifiesto del lote (sistema automático)
        GET /api/v1/batches/{id}/generate_manifest_excel/
        """
        batch = self.get_object()
        
        try:
            # Obtener adapter automáticamente
            adapter = ManifestRegistry.get_adapter_for_entity(batch)
            
            # Generar Excel usando sistema universal
            excel_buffer = UniversalManifestGenerator.generate_excel(batch, adapter)
            
            # Nombre del archivo
            filename = f"manifiesto_lote_{str(batch.id)[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Retornar Excel
            response = HttpResponse(
                excel_buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Error al generar manifiesto Excel: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def export_excel_custom_format(self, request, pk=None):
        """
        Exportar lote a Excel con formato personalizado para transportadora.
        GET /api/v1/batches/{id}/export_excel_custom_format/
        
        Columnas:
        - FECHA: fecha del manifiesto en formato YYYY-MM-DD
        - HORA: hora del manifiesto HH:MM
        - HAWB/REFERENCIA: numero de guia del paquete
        - ESTADO DE GUIA: estado mapeado
        - OBSERVACION: {nombre agencia} - {numero guia agencia}
        - REMESA: vacío
        """
        batch = self.get_object()
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            from io import BytesIO
            from datetime import datetime
            from apps.packages.models import Package
            
            # Obtener todos los paquetes del lote
            packages = Package.objects.filter(pull__batch=batch).select_related(
                'transport_agency', 'pull'
            ).order_by('guide_number')
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Manifiesto"
            
            # Estilos
            header_font = Font(bold=True, size=11)
            normal_font = Font(size=10)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = ['FECHA', 'HORA', 'HAWB/REFERENCIA', 'ESTADO DE GUIA', 'OBSERVACION', 'REMESA']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                ws.column_dimensions[cell.column_letter].width = 20
            
            # Fecha y hora del manifiesto (usar fecha actual o fecha del lote)
            manifest_date = batch.created_at if batch.created_at else datetime.now()
            date_str = manifest_date.strftime('%Y-%m-%d')
            time_str = manifest_date.strftime('%H:%M')
            
            # Función para mapear estado
            def map_status(package_status):
                status_map = {
                    'DEVUELTO': 'DEVOLUCION TRANSPORTADORA',
                    'ENTREGADO': 'ENTREGADA A DESTINATARIO',
                    'EN_TRANSITO': 'ENTREGADA A TRANSPORTADORA',
                    'EN_BODEGA': 'SE RETIRA DEL DESPACHO',
                    'NO_RECEPTADO': 'SE RETIRA DEL DESPACHO',
                    'RETENIDO': 'SE RETIRA DEL DESPACHO',
                }
                return status_map.get(package_status, 'SALE PARA ENTREGA')
            
            # Datos de paquetes
            row = 2
            for pkg in packages:
                # FECHA
                ws.cell(row=row, column=1, value=date_str).font = normal_font
                ws.cell(row=row, column=1).border = border
                
                # HORA
                ws.cell(row=row, column=2, value=time_str).font = normal_font
                ws.cell(row=row, column=2).border = border
                
                # HAWB/REFERENCIA (número de guía del paquete)
                ws.cell(row=row, column=3, value=pkg.guide_number or '').font = normal_font
                ws.cell(row=row, column=3).border = border
                
                # ESTADO DE GUIA
                estado_guia = map_status(pkg.status)
                ws.cell(row=row, column=4, value=estado_guia).font = normal_font
                ws.cell(row=row, column=4).border = border
                
                # OBSERVACION ({nombre agencia} - {numero guia agencia})
                shipping_agency = pkg.get_shipping_agency()
                agency_name = shipping_agency.name if shipping_agency else ''
                agency_guide = pkg.get_shipping_guide_number() or pkg.agency_guide_number or ''
                observacion = f"{agency_name} - {agency_guide}".strip(' -')
                ws.cell(row=row, column=5, value=observacion).font = normal_font
                ws.cell(row=row, column=5).border = border
                
                # REMESA (vacío)
                ws.cell(row=row, column=6, value='').font = normal_font
                ws.cell(row=row, column=6).border = border
                
                row += 1
            
            # Guardar en buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Nombre del archivo
            filename = f"manifiesto_lote_{str(batch.id)[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Retornar Excel
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Error al generar Excel: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def generate_labels(self, request, pk=None):
        """
        Generar PDF con etiquetas de todas las sacas del lote.
        4 etiquetas por página (2x2).
        Cada etiqueta incluye: número de saca (1/5, 2/5...), destino, agencia, 
        guía, cantidad de paquetes y código QR.
        
        GET /api/v1/batches/{id}/generate_labels/
        """
        batch = self.get_object()
        
        try:
            # Generar PDF con etiquetas
            pdf_buffer = BatchLabelsGenerator.generate_pdf(batch)
            
            # Nombre del archivo
            filename = f"etiquetas_lote_{str(batch.id)[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            # Retornar PDF
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Error al generar etiquetas: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def generate_notification_message(self, request, pk=None):
        """
        Generar mensaje de notificación para avisar al usuario que su paquete ha sido enviado.
        GET /api/v1/batches/{id}/generate_notification_message/
        """
        batch = self.get_object()
        
        try:
            from apps.packages.models import Package
            
            # Obtener información del lote
            transport_agency = batch.transport_agency
            guide_number = batch.guide_number
            total_packages = batch.get_total_packages()
            
            if not transport_agency:
                return Response(
                    {'error': 'El lote no tiene agencia de transporte asignada'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if not guide_number:
                return Response(
                    {'error': 'El lote no tiene número de guía asignado'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Determinar URL de seguimiento según la agencia
            tracking_url = None
            agency_name_lower = transport_agency.name.lower()
            if 'yobel' in agency_name_lower:
                tracking_url = 'https://www.yobelscm.biz/ecuador/tracking-clientes-yobel-express-2/'
            
            # Generar mensaje
            message_parts = [
                f"Su paquete ha sido enviado a través de la agencia {transport_agency.name}.",
                f"Número de guía: {guide_number}.",
            ]
            
            if total_packages > 0:
                message_parts.append(f"Se han enviado {total_packages} paquete(s).")
            
            if tracking_url:
                message_parts.append(f"Puede realizar el seguimiento en: {tracking_url}")
            
            message = " ".join(message_parts)
            
            return Response({
                'message': message,
                'agency_name': transport_agency.name,
                'guide_number': guide_number,
                'total_packages': total_packages,
                'tracking_url': tracking_url,
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': f'Error al generar mensaje de notificación: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class DispatchViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar Dispatches
    """
    serializer_class = DispatchSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Retorna el queryset de despachos ordenado"""
        return Dispatch.objects.prefetch_related('pulls', 'packages').order_by('-dispatch_date', '-created_at')
    
    def get_serializer_class(self):
        """Usar serializer detallado para retrieve"""
        if self.action == 'retrieve':
            return DispatchDetailSerializer
        return DispatchSerializer
    
    def create(self, request, *args, **kwargs):
        """
        Crear un nuevo despacho con pulls y packages.
        POST /api/v1/dispatches/
        Body: {
            "dispatch_date": "2025-01-15",
            "status": "PLANIFICADO",
            "notes": "Notas opcionales",
            "pull_ids": ["uuid1", "uuid2", ...],
            "package_ids": ["uuid1", "uuid2", ...]
        }
        """
        pull_ids = request.data.get('pull_ids', [])
        package_ids = request.data.get('package_ids', [])
        
        # Validar que al menos haya pulls o packages
        if not pull_ids and not package_ids:
            return Response(
                {'error': 'Debe proporcionar al menos pull_ids o package_ids'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Crear el dispatch usando el serializer
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        dispatch = serializer.save()
        
        # Asociar pulls si se proporcionaron
        if pull_ids:
            from apps.logistics.models import Pull
            pulls = Pull.objects.filter(id__in=pull_ids)
            if pulls.count() != len(pull_ids):
                return Response(
                    {'error': 'Algunos pull_ids no son válidos'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            dispatch.pulls.set(pulls)
        
        # Asociar packages si se proporcionaron
        if package_ids:
            from apps.packages.models import Package
            packages = Package.objects.filter(id__in=package_ids)
            if packages.count() != len(package_ids):
                return Response(
                    {'error': 'Algunos package_ids no son válidos'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            dispatch.packages.set(packages)
        
        # Retornar el dispatch creado
        response_serializer = self.get_serializer(dispatch)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['get'], url_path='generate-manifest')
    def generate_manifest(self, request, pk=None):
        """
        Genera el manifiesto PDF del despacho (sistema automático).
        GET /api/v1/dispatches/{id}/generate-manifest/
        """
        dispatch = self.get_object()
        
        try:
            # Obtener adapter automáticamente
            adapter = ManifestRegistry.get_adapter_for_entity(dispatch)
            
            # Generar PDF usando sistema universal
            pdf_buffer = UniversalManifestGenerator.generate_pdf(dispatch, adapter)
            
            # Nombre del archivo
            filename = f"manifiesto_despacho_{dispatch.dispatch_date.strftime('%Y%m%d')}_{datetime.now().strftime('%H%M%S')}.pdf"
            
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except Exception as e:
            return Response(
                {'error': f'Error al generar manifiesto: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'], url_path='generate-manifest-excel')
    def generate_manifest_excel(self, request, pk=None):
        """
        Genera el manifiesto Excel del despacho (sistema automático).
        GET /api/v1/dispatches/{id}/generate-manifest-excel/
        """
        dispatch = self.get_object()
        
        try:
            # Obtener adapter automáticamente
            adapter = ManifestRegistry.get_adapter_for_entity(dispatch)
            
            # Generar Excel usando sistema universal
            excel_buffer = UniversalManifestGenerator.generate_excel(dispatch, adapter)
            
            # Nombre del archivo
            filename = f"manifiesto_despacho_{dispatch.dispatch_date.strftime('%Y%m%d')}_{datetime.now().strftime('%H%M%S')}.xlsx"
            
            response = HttpResponse(
                excel_buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except Exception as e:
            return Response(
                {'error': f'Error al generar manifiesto Excel: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'], url_path='export-excel-custom-format')
    def export_excel_custom_format(self, request, pk=None):
        """
        Exportar despacho a Excel con formato personalizado para transportadora.
        GET /api/v1/dispatches/{id}/export-excel-custom-format/?guide_status=ESTADO
        
        Parámetros:
        - guide_status (opcional): Estado de guía a usar para todos los paquetes.
          Valores: DEVOLUCION TRANSPORTADORA, ENTREGADA A DESTINATARIO, 
                   ENTREGADA A TRANSPORTADORA, SALE PARA ENTREGA, SE RETIRA DEL DESPACHO
        
        Columnas:
        - FECHA: fecha del manifiesto en formato YYYY-MM-DD
        - HORA: hora del manifiesto HH:MM
        - HAWB/REFERENCIA: numero de guia del paquete
        - ESTADO DE GUIA: estado seleccionado o mapeado
        - OBSERVACION: {nombre agencia} - {numero guia agencia}
        - REMESA: vacío
        """
        dispatch = self.get_object()
        guide_status = request.query_params.get('guide_status', None)
        
        # Estados válidos
        valid_statuses = [
            'DEVOLUCION TRANSPORTADORA',
            'ENTREGADA A DESTINATARIO',
            'ENTREGADA A TRANSPORTADORA',
            'SALE PARA ENTREGA',
            'SE RETIRA DEL DESPACHO'
        ]
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            from io import BytesIO
            from datetime import datetime
            from apps.packages.models import Package
            
            # Obtener todos los paquetes del despacho
            # Paquetes de pulls
            packages_from_pulls = Package.objects.filter(
                pull__in=dispatch.pulls.all()
            ).select_related('transport_agency', 'pull', 'pull__batch')
            
            # Paquetes individuales
            packages_individual = dispatch.packages.select_related('transport_agency')
            
            # Combinar y ordenar
            all_packages = list(packages_from_pulls) + list(packages_individual)
            all_packages.sort(key=lambda p: p.guide_number or '')
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Manifiesto"
            
            # Estilos
            header_font = Font(bold=True, size=11)
            normal_font = Font(size=10)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = ['FECHA', 'HORA', 'HAWB/REFERENCIA', 'ESTADO DE GUIA', 'OBSERVACION', 'REMESA']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                ws.column_dimensions[cell.column_letter].width = 20
            
            # Fecha y hora del manifiesto
            manifest_date = dispatch.dispatch_date if dispatch.dispatch_date else datetime.now().date()
            manifest_datetime = datetime.combine(manifest_date, datetime.now().time())
            date_str = manifest_datetime.strftime('%Y-%m-%d')
            time_str = manifest_datetime.strftime('%H:%M')
            
            # Función para obtener estado
            def get_guide_status(pkg):
                if guide_status and guide_status in valid_statuses:
                    return guide_status
                # Mapeo automático si no se especifica
                status_map = {
                    'DEVUELTO': 'DEVOLUCION TRANSPORTADORA',
                    'ENTREGADO': 'ENTREGADA A DESTINATARIO',
                    'EN_TRANSITO': 'ENTREGADA A TRANSPORTADORA',
                    'EN_BODEGA': 'SE RETIRA DEL DESPACHO',
                    'NO_RECEPTADO': 'SE RETIRA DEL DESPACHO',
                    'RETENIDO': 'SE RETIRA DEL DESPACHO',
                }
                return status_map.get(pkg.status, 'SALE PARA ENTREGA')
            
            # Datos de paquetes
            row = 2
            for pkg in all_packages:
                # FECHA
                ws.cell(row=row, column=1, value=date_str).font = normal_font
                ws.cell(row=row, column=1).border = border
                
                # HORA
                ws.cell(row=row, column=2, value=time_str).font = normal_font
                ws.cell(row=row, column=2).border = border
                
                # HAWB/REFERENCIA (número de guía del paquete)
                ws.cell(row=row, column=3, value=pkg.guide_number or '').font = normal_font
                ws.cell(row=row, column=3).border = border
                
                # ESTADO DE GUIA
                estado_guia = get_guide_status(pkg)
                ws.cell(row=row, column=4, value=estado_guia).font = normal_font
                ws.cell(row=row, column=4).border = border
                
                # OBSERVACION ({nombre agencia} - {numero guia agencia})
                shipping_agency = pkg.get_shipping_agency()
                agency_name = shipping_agency.name if shipping_agency else ''
                agency_guide = pkg.get_shipping_guide_number() or pkg.agency_guide_number or ''
                observacion = f"{agency_name} - {agency_guide}".strip(' -')
                ws.cell(row=row, column=5, value=observacion).font = normal_font
                ws.cell(row=row, column=5).border = border
                
                # REMESA (vacío)
                ws.cell(row=row, column=6, value='').font = normal_font
                ws.cell(row=row, column=6).border = border
                
                row += 1
            
            # Guardar en buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Nombre del archivo
            filename = f"manifiesto_despacho_{dispatch.dispatch_date.strftime('%Y%m%d')}_{datetime.now().strftime('%H%M%S')}.xlsx"
            
            # Retornar Excel
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Error al generar Excel: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='by-date')
    def get_by_date(self, request):
        """
        Obtiene despachos por fecha.
        GET /api/v1/dispatches/by-date/?date=2025-01-15
        """
        date_str = request.query_params.get('date', None)
        if not date_str:
            return Response(
                {'error': 'El parámetro "date" es requerido (formato: YYYY-MM-DD)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from datetime import datetime as dt
            dispatch_date = dt.strptime(date_str, '%Y-%m-%d').date()
            dispatches = Dispatch.objects.filter(dispatch_date=dispatch_date)
            serializer = self.get_serializer(dispatches, many=True)
            return Response(serializer.data)
        except ValueError:
            return Response(
                {'error': 'Formato de fecha inválido. Use YYYY-MM-DD'},
                status=status.HTTP_400_BAD_REQUEST
            )


class ExcelMapperViewSet(viewsets.ViewSet):
    """
    ViewSet para mapear Excel de entrada a formato de salida personalizado
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    @action(detail=False, methods=['post'], url_path='read-columns')
    def read_columns(self, request):
        """
        Lee las columnas de un archivo Excel subido
        POST /api/v1/excel-mapper/read-columns/
        Body: multipart/form-data con campo 'file'
        """
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No se proporcionó ningún archivo'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        
        try:
            import openpyxl
            from io import BytesIO
            
            # Leer Excel
            wb = openpyxl.load_workbook(BytesIO(file.read()), read_only=True)
            ws = wb.active
            
            # Obtener columnas de la primera fila
            columns = []
            for cell in ws[1]:
                if cell.value:
                    columns.append(str(cell.value).strip())
            
            # Obtener algunas filas de ejemplo (máximo 5)
            sample_rows = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), start=2):
                if any(row):  # Si la fila no está completamente vacía
                    sample_rows.append({
                        'row': row_idx,
                        'data': [str(cell) if cell is not None else '' for cell in row[:len(columns)]]
                    })
            
            return Response({
                'columns': columns,
                'sample_rows': sample_rows,
                'total_rows': ws.max_row - 1  # Excluyendo header
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Error al leer el archivo Excel: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'], url_path='generate-output')
    def generate_output(self, request):
        """
        Genera un Excel de salida con el formato personalizado basado en el mapeo
        POST /api/v1/excel-mapper/generate-output/
        Body: multipart/form-data con:
            - 'file': archivo Excel
            - 'mapping': JSON string con el mapeo
        """
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No se proporcionó ningún archivo'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        mapping_str = request.data.get('mapping', '{}')
        
        try:
            import json
            mapping = json.loads(mapping_str) if isinstance(mapping_str, str) else mapping_str
        except json.JSONDecodeError:
            return Response(
                {'error': 'Formato de mapeo inválido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
            from io import BytesIO
            from datetime import datetime
            
            # Leer Excel de entrada
            wb_input = openpyxl.load_workbook(BytesIO(file.read()), read_only=True)
            ws_input = wb_input.active
            
            # Leer headers
            headers = []
            for cell in ws_input[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append('')
            
            # Crear Excel de salida
            wb_output = openpyxl.Workbook()
            ws_output = wb_output.active
            ws_output.title = "Manifiesto"
            
            # Estilos
            header_font = Font(bold=True, size=11)
            normal_font = Font(size=10)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers de salida
            output_headers = ['FECHA', 'HORA', 'HAWB/REFERENCIA', 'ESTADO DE GUIA', 'OBSERVACION', 'REMESA']
            for col, header in enumerate(output_headers, 1):
                cell = ws_output.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                ws_output.column_dimensions[cell.column_letter].width = 20
            
            # Fecha y hora actuales
            now = datetime.now()
            date_str = now.strftime('%Y-%m-%d')
            time_str = now.strftime('%H:%M')
            
            # Función para obtener valor de celda
            def get_cell_value(row_data, col_index):
                if col_index is None or col_index == '' or col_index == 'none':
                    return ''
                try:
                    idx = int(col_index)
                    if 0 <= idx < len(row_data):
                        val = row_data[idx]
                        return str(val) if val is not None and str(val).strip() else ''
                except (ValueError, IndexError, TypeError):
                    pass
                return ''
            
            # Función para mapear estado
            def get_estado_guia(row_data, estado_mapping):
                if not estado_mapping or estado_mapping == '':
                    return ''
                if estado_mapping == 'today' or estado_mapping == 'now':
                    return ''
                if estado_mapping.startswith('col_'):
                    col_idx = estado_mapping.replace('col_', '')
                    return get_cell_value(row_data, col_idx)
                # Si es un valor fijo
                valid_statuses = [
                    'DEVOLUCION TRANSPORTADORA',
                    'ENTREGADA A DESTINATARIO',
                    'ENTREGADA A TRANSPORTADORA',
                    'SALE PARA ENTREGA',
                    'SE RETIRA DEL DESPACHO'
                ]
                if estado_mapping in valid_statuses:
                    return estado_mapping
                return ''
            
            # Procesar filas
            row_output = 2
            for row_idx, row in enumerate(ws_input.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Saltar filas vacías
                    continue
                
                row_data = [str(cell) if cell is not None else '' for cell in row[:len(headers)]]
                
                # FECHA
                fecha_mapping = mapping.get('fecha', 'today')
                if fecha_mapping == 'today':
                    fecha_value = date_str
                else:
                    fecha_value = get_cell_value(row_data, fecha_mapping.replace('col_', ''))
                ws_output.cell(row=row_output, column=1, value=fecha_value).font = normal_font
                ws_output.cell(row=row_output, column=1).border = border
                
                # HORA
                hora_mapping = mapping.get('hora', 'now')
                if hora_mapping == 'now':
                    hora_value = time_str
                else:
                    hora_value = get_cell_value(row_data, hora_mapping.replace('col_', ''))
                ws_output.cell(row=row_output, column=2, value=hora_value).font = normal_font
                ws_output.cell(row=row_output, column=2).border = border
                
                # HAWB/REFERENCIA
                hawb_mapping = mapping.get('hawb', '')
                hawb_value = get_cell_value(row_data, hawb_mapping.replace('col_', ''))
                ws_output.cell(row=row_output, column=3, value=hawb_value).font = normal_font
                ws_output.cell(row=row_output, column=3).border = border
                
                # ESTADO DE GUIA
                estado_mapping = mapping.get('estado', '')
                estado_value = get_estado_guia(row_data, estado_mapping)
                ws_output.cell(row=row_output, column=4, value=estado_value).font = normal_font
                ws_output.cell(row=row_output, column=4).border = border
                
                # OBSERVACION
                agency = mapping.get('observacion_agency', '')
                col1 = get_cell_value(row_data, mapping.get('observacion_col1', '').replace('col_', ''))
                col2 = get_cell_value(row_data, mapping.get('observacion_col2', '').replace('col_', ''))
                observacion_parts = [p for p in [agency, col1, col2] if p and p.strip()]
                observacion_value = ' - '.join(observacion_parts)
                ws_output.cell(row=row_output, column=5, value=observacion_value).font = normal_font
                ws_output.cell(row=row_output, column=5).border = border
                
                # REMESA (vacío)
                ws_output.cell(row=row_output, column=6, value='').font = normal_font
                ws_output.cell(row=row_output, column=6).border = border
                
                row_output += 1
            
            # Guardar en buffer
            buffer = BytesIO()
            wb_output.save(buffer)
            buffer.seek(0)
            
            # Nombre del archivo
            filename = f"manifiesto_mapeado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Retornar Excel
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Error al generar Excel: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

