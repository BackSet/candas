"""
Servicio para generar manifiestos de lotes.
Formato A4 horizontal, formal, sin colores.
"""
from io import BytesIO
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from datetime import datetime
from apps.shared.services.manifest_template import BaseManifestGenerator


class BatchManifestGenerator:
    """Generador de manifiestos para lotes."""
    
    @staticmethod
    def generate_pdf(batch):
        """
        Genera un PDF con el manifiesto del lote.
        Formato A4 horizontal, formal, sin colores.
        Estructura: Información del lote, luego cada saca con sus paquetes debajo.
        """
        doc, buffer = BaseManifestGenerator.create_document()
        elements = []
        styles = getSampleStyleSheet()
        
        # Header de la empresa
        elements.append(BaseManifestGenerator.get_company_header())
        elements.append(Spacer(1, 3*mm))
        
        # Título
        title_style = BaseManifestGenerator.get_title_style()
        elements.append(Paragraph("MANIFIESTO DE LOTE", title_style))
        elements.append(Spacer(1, 3*mm))
        
        # Información del lote (sin tabla, solo párrafos)
        normal_style = BaseManifestGenerator.get_normal_style()
        
        # Crear párrafos separados por saltos de línea
        info_lines = [
            f"Destino: {batch.destiny}",
            f"Agencia de Transporte: {batch.transport_agency.name if batch.transport_agency else 'Sin asignar'}",
            f"Número de Guía de Transporte: {batch.guide_number or 'Sin guía'}",
            f"Fecha: {BaseManifestGenerator.format_datetime_now()}",
            f"Total Sacas: {str(batch.get_pull_count())}",
            f"Total Paquetes: {str(batch.get_total_packages())}",
        ]
        
        # Agregar cada línea como un párrafo usando map
        elements.extend(map(lambda line: Paragraph(line, normal_style), info_lines))
        
        elements.append(Spacer(1, 5*mm))
        
        # Nota general sobre las notas
        normal_style = BaseManifestGenerator.get_normal_style()
        warning_style = ParagraphStyle(
            'WarningStyle',
            parent=normal_style,
            fontSize=8,
            textColor=colors.black,
            fontName='Helvetica-Bold',
            spaceAfter=5,
            spaceBefore=5
        )
        warning_text = "IMPORTANTE: Preste atención a la columna de NOTAS, ya que puede contener información importante de envío. Si no hay nota, utilice la información proporcionada en las demás celdas."
        elements.append(Paragraph(warning_text, warning_style))
        elements.append(Spacer(1, 3*mm))
        
        # Obtener sacas ordenadas
        pulls = batch.pulls.all().order_by('created_at')
        
        if pulls.exists():
            from apps.packages.models import Package
            
            # Para cada saca, mostrar encabezado y sus paquetes
            for idx, pull in enumerate(pulls, 1):
                # Encabezado de saca
                heading_style = BaseManifestGenerator.get_heading_style()
                saca_title = f"SACA {idx} de {pulls.count()}"
                elements.append(Paragraph(saca_title, heading_style))
                
                # Información de la saca (solo tamaño y cantidad, el resto ya está en el lote)
                # Sin tabla, solo párrafos
                normal_style = BaseManifestGenerator.get_normal_style()
                saca_info_lines = [
                    f"Tamaño: {pull.get_size_display()}",
                    f"Cantidad de Paquetes: {str(pull.packages.count())}",
                ]
                
                # Agregar cada línea como un párrafo usando map
                elements.extend(map(lambda line: Paragraph(line, normal_style), saca_info_lines))
                
                elements.append(Spacer(1, 3*mm))
                
                # Paquetes de esta saca
                packages = pull.packages.all().order_by('guide_number')
                
                if packages.exists():
                    # Encabezados de tabla (sin Nota, ahora va debajo)
                    pkg_headers = ['#', 'Número de Guía', 'Destinatario', 'Dirección', 'Ciudad', 'Provincia', 'Teléfono', 'Firma']
                    
                    # Construir datos de tabla con nota expandida debajo de cada paquete
                    table_data = []
                    normal_style = BaseManifestGenerator.get_normal_style()
                    note_style = ParagraphStyle(
                        'NoteStyle',
                        parent=normal_style,
                        fontSize=7,
                        textColor=colors.black,
                        fontName='Helvetica-Bold',
                        leftIndent=5,
                        rightIndent=5
                    )
                    
                    # Convertir headers a Paragraphs
                    bold_style = ParagraphStyle(
                        'FormalBold',
                        parent=normal_style,
                        fontName='Helvetica-Bold',
                        fontSize=8
                    )
                    wrapped_headers = [BaseManifestGenerator._wrap_text(h, bold_style) for h in pkg_headers]
                    table_data.append(wrapped_headers)
                    
                    for pkg_idx, package in enumerate(packages, 1):
                        # Fila de datos del paquete
                        data_style = ParagraphStyle(
                            'FormalData',
                            parent=normal_style,
                            fontSize=7
                        )
                        pkg_row = [
                            BaseManifestGenerator._wrap_text(str(pkg_idx), data_style),
                            BaseManifestGenerator._wrap_text(package.guide_number, data_style),
                            BaseManifestGenerator._wrap_text(package.name, data_style),
                            BaseManifestGenerator._wrap_text(package.address, data_style),
                            BaseManifestGenerator._wrap_text(package.city, data_style),
                            BaseManifestGenerator._wrap_text(package.province, data_style),
                            BaseManifestGenerator._wrap_text(package.phone_number or '-', data_style),
                            BaseManifestGenerator._wrap_text('', data_style),  # Celda de firma vacía
                        ]
                        table_data.append(pkg_row)
                        
                        # Fila expandida con la nota (si existe)
                        note_text = package.notes if package.notes else 'Sin nota - usar información de las celdas'
                        note_cell = BaseManifestGenerator._wrap_text(f"NOTA: {note_text}", note_style)
                        # Crear fila con nota expandida (ocupa todas las columnas excepto Firma)
                        # La columna Firma se deja vacía porque se unirá verticalmente con la fila anterior
                        note_row = [note_cell] + [BaseManifestGenerator._wrap_text('', data_style)] * (len(pkg_headers) - 2) + [BaseManifestGenerator._wrap_text('', data_style)]
                        table_data.append(note_row)
                    
                    # Calcular ancho disponible: A4 horizontal (297mm) - márgenes (10mm * 2) = 277mm
                    available_width = 277 * mm
                    # Anchos de columnas proporcionales para ocupar todo el ancho disponible
                    col_widths = [
                        15*mm,      # #
                        35*mm,      # Número de Guía
                        40*mm,      # Destinatario
                        50*mm,      # Dirección
                        30*mm,      # Ciudad
                        30*mm,      # Provincia
                        30*mm,      # Teléfono
                        47*mm       # Firma (el resto del espacio)
                    ]
                    
                    # Ajustar para que sumen exactamente el ancho disponible
                    total_width = sum(col_widths)
                    if total_width != available_width:
                        # Ajustar proporcionalmente
                        ratio = available_width / total_width
                        col_widths = [w * ratio for w in col_widths]
                    
                    # Crear tabla
                    table = Table(table_data, colWidths=col_widths)
                    
                    # Estilo de tabla con padding reducido
                    table_style = [
                        # Encabezado
                        ('BACKGROUND', (0, 0), (-1, 0), colors.white),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 8),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 3),
                        ('TOPPADDING', (0, 0), (-1, 0), 3),
                        ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.black),
                        # Datos
                        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 1), (-1, -1), 7),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('LEFTPADDING', (0, 0), (-1, -1), 2),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                        ('TOPPADDING', (0, 1), (-1, -1), 2),
                        ('BOTTOMPADDING', (0, 1), (-1, -1), 2),
                    ]
                    
                    # Aplicar estilo especial a las filas de nota (filas pares después del header)
                    # Unificar visualmente la fila de nota con la fila del paquete
                    for row_idx in range(2, len(table_data), 2):  # Solo filas de nota (empezando en índice 2)
                        data_row_idx = row_idx - 1  # Fila de datos del paquete (anterior a la nota)
                        # Nota ocupa todas las columnas excepto Firma
                        table_style.append(('SPAN', (0, row_idx), (-2, row_idx)))  # Nota ocupa todas las columnas excepto la última (Firma)
                        table_style.append(('BACKGROUND', (0, row_idx), (-2, row_idx), colors.HexColor('#f5f5f5')))
                        table_style.append(('FONTNAME', (0, row_idx), (-2, row_idx), 'Helvetica-Bold'))
                        # Unir la columna de Firma verticalmente (fila del paquete + fila de nota)
                        firma_col_idx = len(pkg_headers) - 1  # Índice de la última columna (Firma)
                        table_style.append(('SPAN', (firma_col_idx, data_row_idx), (firma_col_idx, row_idx)))  # Unir Firma verticalmente
                        # Eliminar bordes horizontales entre filas para unificar visualmente
                        # Eliminar borde inferior de la fila de datos (excepto en la columna Firma que ya está unida)
                        table_style.append(('LINEBELOW', (0, data_row_idx), (-2, data_row_idx), 0, colors.white))
                        # Eliminar borde superior de la fila de nota (excepto en la columna Firma que ya está unida)
                        table_style.append(('LINEABOVE', (0, row_idx), (-2, row_idx), 0, colors.white))
                        # Mantener padding reducido en la fila de nota
                        table_style.append(('TOPPADDING', (0, row_idx), (-2, row_idx), 1))
                        table_style.append(('BOTTOMPADDING', (0, row_idx), (-2, row_idx), 2))
                    
                    table.setStyle(TableStyle(table_style))
                    elements.append(table)
                else:
                    normal_style = BaseManifestGenerator.get_normal_style()
                    elements.append(Paragraph("No hay paquetes en esta saca", normal_style))
                
                elements.append(Spacer(1, 5*mm))
                
                # Salto de página si no es la última saca
                if idx < pulls.count():
                    elements.append(PageBreak())
        else:
            normal_style = BaseManifestGenerator.get_normal_style()
            elements.append(Paragraph("No hay sacas en este lote", normal_style))
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        
        return buffer
    
    @staticmethod
    def generate_excel(batch):
        """
        Genera un Excel con el manifiesto del lote.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Manifiesto"
        
        # Estilos formales (sin colores)
        header_font = Font(bold=True, size=12)
        normal_font = Font(size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:G1')
        ws['A1'] = 'MANIFIESTO DE LOTE'
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Información del lote
        row = 3
        ws[f'A{row}'] = 'INFORMACIÓN DEL LOTE'
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        
        row += 1
        info_data = [
            ('Destino:', batch.destiny),
            ('Agencia de Transporte:', batch.transport_agency.name if batch.transport_agency else 'Sin asignar'),
            ('Número de Guía de Transporte:', batch.guide_number or 'Sin guía'),
            ('Fecha:', datetime.now().strftime('%d/%m/%Y %H:%M')),
            ('Total Sacas:', batch.get_pull_count()),
            ('Total Paquetes:', batch.get_total_packages()),
        ]
        
        # Usar zip para iteración paralela
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
        
        # Sacas y paquetes
        pulls = batch.pulls.all().order_by('created_at')
        from apps.packages.models import Package
        
        for pull_idx, pull in enumerate(pulls, 1):
            row += 2
            ws[f'A{row}'] = f'SACA {pull_idx} de {pulls.count()}'
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            
            # Información de la saca (solo tamaño y cantidad, el resto ya está en el lote)
            row += 1
            saca_info = [
                ('Tamaño:', pull.get_size_display()),
                ('Cantidad de Paquetes:', str(pull.packages.count())),
            ]
            
            # Usar zip para iteración paralela
            for label, value in saca_info:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                row += 1
            
            # Paquetes
            packages = pull.packages.all().order_by('guide_number')
            if packages.exists():
                row += 1
                pkg_headers = ['#', 'Número de Guía', 'Destinatario', 'Dirección', 'Ciudad', 'Provincia', 'Teléfono', 'Nota']
                for col, header in enumerate(pkg_headers, 1):
                    cell = ws.cell(row, col, header)
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border
                
                row += 1
                for pkg_idx, package in enumerate(packages, 1):
                    ws.cell(row, 1, pkg_idx).border = border
                    ws.cell(row, 2, package.guide_number).border = border
                    ws.cell(row, 3, package.name).border = border
                    ws.cell(row, 4, package.address).border = border
                    ws.cell(row, 5, package.city).border = border
                    ws.cell(row, 6, package.province).border = border
                    ws.cell(row, 7, package.phone_number or '-').border = border
                    # Nota resaltada
                    note_cell = ws.cell(row, 8, package.notes or '-')
                    note_cell.border = border
                    note_cell.font = Font(bold=True)
                    note_cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                    row += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 28
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 30
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
