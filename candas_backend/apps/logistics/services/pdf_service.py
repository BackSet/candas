"""
Servicio para generación de PDFs
"""
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.pdfgen import canvas
from datetime import datetime
from PIL import Image as PILImage
import io
import qrcode
from reportlab.lib.utils import ImageReader
from .qr_service import QRService
from apps.shared.services.manifest_template import BaseManifestGenerator


class PDFService:
    """Servicio para generar PDFs de sacas"""
    
    @staticmethod
    def generate_pull_manifest(pull):
        """
        Genera PDF con manifiesto de paquetes de una saca.
        Formato A4 horizontal, formal, sin colores.
        
        Args:
            pull: Instancia del modelo Pull
            
        Returns:
            BytesIO con el PDF generado
        """
        doc, buffer = BaseManifestGenerator.create_document()
        elements = []
        
        # Header de la empresa
        elements.append(BaseManifestGenerator.get_company_header())
        elements.append(Spacer(1, 3*mm))
        
        # Título
        title_style = BaseManifestGenerator.get_title_style()
        elements.append(Paragraph("MANIFIESTO DE SACA", title_style))
        elements.append(Spacer(1, 3*mm))
        
        # Información de la saca (sin tabla, solo párrafos)
        effective_agency = pull.get_effective_agency()
        effective_guide = pull.get_effective_guide_number()
        effective_destiny = pull.get_effective_destiny()
        
        normal_style = BaseManifestGenerator.get_normal_style()
        bold_style = ParagraphStyle(
            'FormalBold',
            parent=normal_style,
            fontName='Helvetica-Bold',
            fontSize=8
        )
        
        # Crear párrafos separados por saltos de línea
        info_lines = [
            f"Destino: {effective_destiny}",
            f"Agencia de Transporte: {effective_agency.name if effective_agency else 'Sin asignar'}",
            f"Número de Guía de Transporte: {effective_guide or 'Sin guía'}",
            f"Tamaño: {pull.get_size_display()}",
            f"Fecha: {BaseManifestGenerator.format_datetime_now()}",
            f"Cantidad de Paquetes: {str(pull.packages.count())}",
        ]
        
        # Agregar información del lote si existe
        if pull.batch:
            info_lines.append(f"Lote: {str(pull.batch.id)[:8].upper()} - {pull.batch.destiny}")
        
        # Agregar cada línea como un párrafo
        for line in info_lines:
            elements.append(Paragraph(line, normal_style))
        
        elements.append(Spacer(1, 5*mm))
        
        # Nota general sobre las notas
        normal_style = BaseManifestGenerator.get_normal_style()
        warning_style = ParagraphStyle(
            'WarningStyle',
            parent=normal_style,
            fontSize=8,
            textColor=colors.black,
            fontName='Helvetica-Bold',
            spaceAfter=5,
            spaceBefore=5
        )
        warning_text = "IMPORTANTE: Preste atención a la columna de NOTAS, ya que puede contener información importante de envío. Si no hay nota, utilice la información proporcionada en las demás celdas."
        elements.append(Paragraph(warning_text, warning_style))
        elements.append(Spacer(1, 3*mm))
        
        # Lista de paquetes
        heading_style = BaseManifestGenerator.get_heading_style()
        elements.append(Paragraph("LISTA DE PAQUETES", heading_style))
        elements.append(Spacer(1, 3*mm))
        
        packages = pull.packages.all().order_by('guide_number')
        
        if packages.exists():
            # Encabezados de tabla (sin Nota, ahora va debajo)
            pkg_headers = ['#', 'Número de Guía', 'Destinatario', 'Dirección', 'Ciudad', 'Provincia', 'Teléfono', 'Firma']
            
            # Construir datos de tabla con nota expandida debajo de cada paquete
            table_data = []
            normal_style = BaseManifestGenerator.get_normal_style()
            note_style = ParagraphStyle(
                'NoteStyle',
                parent=normal_style,
                fontSize=7,
                textColor=colors.black,
                fontName='Helvetica-Bold',
                leftIndent=5,
                rightIndent=5
            )
            
            # Convertir headers a Paragraphs
            bold_style = ParagraphStyle(
                'FormalBold',
                parent=normal_style,
                fontName='Helvetica-Bold',
                fontSize=8
            )
            wrapped_headers = [BaseManifestGenerator._wrap_text(h, bold_style) for h in pkg_headers]
            table_data.append(wrapped_headers)
            
            for idx, package in enumerate(packages, 1):
                # Fila de datos del paquete
                data_style = ParagraphStyle(
                    'FormalData',
                    parent=normal_style,
                    fontSize=7
                )
                pkg_row = [
                    BaseManifestGenerator._wrap_text(str(idx), data_style),
                    BaseManifestGenerator._wrap_text(package.guide_number, data_style),
                    BaseManifestGenerator._wrap_text(package.name, data_style),
                    BaseManifestGenerator._wrap_text(package.address, data_style),
                    BaseManifestGenerator._wrap_text(package.city, data_style),
                    BaseManifestGenerator._wrap_text(package.province, data_style),
                    BaseManifestGenerator._wrap_text(package.phone_number or '-', data_style),
                    BaseManifestGenerator._wrap_text('', data_style),  # Celda de firma vacía
                ]
                table_data.append(pkg_row)
                
                # Fila expandida con la nota (si existe)
                note_text = package.notes if package.notes else 'Sin nota - usar información de las celdas'
                note_cell = BaseManifestGenerator._wrap_text(f"NOTA: {note_text}", note_style)
                # Crear fila con nota expandida (ocupa todas las columnas excepto Firma)
                # La columna Firma se deja vacía porque se unirá verticalmente con la fila anterior
                note_row = [note_cell] + [BaseManifestGenerator._wrap_text('', data_style)] * (len(pkg_headers) - 2) + [BaseManifestGenerator._wrap_text('', data_style)]
                table_data.append(note_row)
            
            # Calcular ancho disponible: A4 horizontal (297mm) - márgenes (10mm * 2) = 277mm
            available_width = 277 * mm
            # Anchos de columnas proporcionales para ocupar todo el ancho disponible
            col_widths = [
                15*mm,      # #
                35*mm,      # Número de Guía
                40*mm,      # Destinatario
                50*mm,      # Dirección
                30*mm,      # Ciudad
                30*mm,      # Provincia
                30*mm,      # Teléfono
                47*mm       # Firma (el resto del espacio)
            ]
            
            # Ajustar para que sumen exactamente el ancho disponible
            total_width = sum(col_widths)
            if total_width != available_width:
                # Ajustar proporcionalmente
                ratio = available_width / total_width
                col_widths = [w * ratio for w in col_widths]
            
            # Crear tabla
            table = Table(table_data, colWidths=col_widths)
            
            # Estilo de tabla con padding reducido
            table_style = [
                # Encabezado
                ('BACKGROUND', (0, 0), (-1, 0), colors.white),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 3),
                ('TOPPADDING', (0, 0), (-1, 0), 3),
                ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.black),
                # Datos
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 2),
                ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                ('TOPPADDING', (0, 1), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 2),
            ]
            
            # Aplicar estilo especial a las filas de nota (filas pares después del header)
            # Unificar visualmente la fila de nota con la fila del paquete
            for row_idx in range(2, len(table_data), 2):  # Solo filas de nota (empezando en índice 2)
                data_row_idx = row_idx - 1  # Fila de datos del paquete (anterior a la nota)
                # Nota ocupa todas las columnas excepto Firma
                table_style.append(('SPAN', (0, row_idx), (-2, row_idx)))  # Nota ocupa todas las columnas excepto la última (Firma)
                table_style.append(('BACKGROUND', (0, row_idx), (-2, row_idx), colors.HexColor('#f5f5f5')))
                table_style.append(('FONTNAME', (0, row_idx), (-2, row_idx), 'Helvetica-Bold'))
                # Unir la columna de Firma verticalmente (fila del paquete + fila de nota)
                firma_col_idx = len(pkg_headers) - 1  # Índice de la última columna (Firma)
                table_style.append(('SPAN', (firma_col_idx, data_row_idx), (firma_col_idx, row_idx)))  # Unir Firma verticalmente
                # Eliminar bordes horizontales entre filas para unificar visualmente
                # Eliminar borde inferior de la fila de datos (excepto en la columna Firma que ya está unida)
                table_style.append(('LINEBELOW', (0, data_row_idx), (-2, data_row_idx), 0, colors.white))
                # Eliminar borde superior de la fila de nota (excepto en la columna Firma que ya está unida)
                table_style.append(('LINEABOVE', (0, row_idx), (-2, row_idx), 0, colors.white))
                # Mantener padding reducido en la fila de nota
                table_style.append(('TOPPADDING', (0, row_idx), (-2, row_idx), 1))
                table_style.append(('BOTTOMPADDING', (0, row_idx), (-2, row_idx), 2))
            
            table.setStyle(TableStyle(table_style))
            elements.append(table)
        else:
            normal_style = BaseManifestGenerator.get_normal_style()
            elements.append(Paragraph("No hay paquetes en esta saca", normal_style))
        
        # Generar PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def generate_pull_manifest_excel(pull):
        """
        Genera un Excel con el manifiesto de la saca.
        """
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from datetime import datetime
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Manifiesto"
        
        # Estilos formales (sin colores)
        header_font = Font(bold=True, size=12)
        normal_font = Font(size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:B1')
        ws['A1'] = 'MANIFIESTO DE SACA'
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Información de la saca
        effective_agency = pull.get_effective_agency()
        effective_guide = pull.get_effective_guide_number()
        effective_destiny = pull.get_effective_destiny()
        
        row = 3
        ws[f'A{row}'] = 'INFORMACIÓN DE LA SACA'
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        
        row += 1
        saca_info = [
            ('Destino:', effective_destiny),
            ('Agencia de Transporte:', effective_agency.name if effective_agency else 'Sin asignar'),
            ('Número de Guía de Transporte:', effective_guide or 'Sin guía'),
            ('Tamaño:', pull.get_size_display()),
            ('Fecha:', datetime.now().strftime('%d/%m/%Y %H:%M')),
            ('Cantidad de Paquetes:', str(pull.packages.count())),
        ]
        
        # Agregar información del lote si existe
        if pull.batch:
            saca_info.append(('Lote:', f"{str(pull.batch.id)[:8].upper()} - {pull.batch.destiny}"))
        
        for label, value in saca_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
        
        # Lista de paquetes
        packages = pull.packages.all().order_by('guide_number')
        if packages.exists():
            row += 1
            ws[f'A{row}'] = 'LISTA DE PAQUETES'
            ws.merge_cells(f'A{row}:H{row}')
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            
            row += 1
            pkg_headers = ['#', 'Número de Guía', 'Destinatario', 'Dirección', 'Ciudad', 'Provincia', 'Teléfono', 'Nota']
            for col, header in enumerate(pkg_headers, 1):
                cell = ws.cell(row, col, header)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            row += 1
            for pkg_idx, package in enumerate(packages, 1):
                ws.cell(row, 1, pkg_idx).border = border
                ws.cell(row, 2, package.guide_number).border = border
                ws.cell(row, 3, package.name).border = border
                ws.cell(row, 4, package.address).border = border
                ws.cell(row, 5, package.city).border = border
                ws.cell(row, 6, package.province).border = border
                ws.cell(row, 7, package.phone_number or '-').border = border
                # Nota resaltada
                note_cell = ws.cell(row, 8, package.notes or '-')
                note_cell.border = border
                if package.notes:
                    note_cell.font = Font(bold=True)
                    note_cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                row += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 30
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    @staticmethod
    def generate_pull_label(pull):
        """
        Genera PDF con etiqueta de saca usando formato estándar (4 etiquetas por página, 2x2).
        Mismo formato que etiquetas de paquetes y lotes.
        
        Args:
            pull: Instancia del modelo Pull
            
        Returns:
            BytesIO con el PDF generado
        """
        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=letter)
        page_width, page_height = letter
        
        # Configuración de etiquetas (igual que PackageLabelsGenerator y BatchLabelsGenerator)
        LABEL_WIDTH = 4 * inch
        LABEL_HEIGHT = 5 * inch
        MARGIN_X = 0.25 * inch
        MARGIN_Y = 0.25 * inch
        
        # Dibujar etiqueta (1/1 ya que es una sola saca)
        x = MARGIN_X
        y = page_height - MARGIN_Y - LABEL_HEIGHT
        
        PDFService._draw_pull_label(c, pull, x, y, LABEL_WIDTH, LABEL_HEIGHT)
        
        c.showPage()
        c.save()
        buffer.seek(0)
        
        return buffer
    
    @staticmethod
    def _draw_pull_label(canvas_obj, pull, x, y, label_width, label_height):
        """
        Dibuja una etiqueta individual para una saca.
        Formato igual a etiquetas de paquetes y lotes.
        
        Args:
            canvas_obj: Objeto Canvas de ReportLab
            pull: Instancia del modelo Pull
            x, y: Posición de la etiqueta
            label_width: Ancho de la etiqueta
            label_height: Alto de la etiqueta
        """
        # Borde de la etiqueta
        canvas_obj.setStrokeColor(colors.black)
        canvas_obj.setLineWidth(2)
        canvas_obj.rect(x, y, label_width, label_height)
        
        # Padding interno
        padding = 0.2 * inch
        content_x = x + padding
        content_y = y + label_height - padding
        content_width = label_width - 2 * padding
        
        # Numeración (grande y destacado) - formato 1/1
        canvas_obj.setFont("Helvetica-Bold", 32)
        canvas_obj.setFillColor(colors.black)
        numeracion = "1/1"
        canvas_obj.drawCentredString(
            x + label_width/2,
            content_y - 0.5*inch,
            numeracion
        )
        
        # Línea separadora
        canvas_obj.setLineWidth(1)
        canvas_obj.line(
            content_x,
            content_y - 0.8*inch,
            content_x + content_width,
            content_y - 0.8*inch
        )
        
        # Información de la saca
        current_y = content_y - 1.2*inch
        
        # Obtener datos efectivos (heredados del batch si existen)
        effective_destiny = pull.get_effective_destiny()
        effective_agency = pull.get_effective_agency()
        effective_guide = pull.get_effective_guide_number()
        
        # Destino
        canvas_obj.setFont("Helvetica-Bold", 10)
        canvas_obj.drawString(content_x, current_y, "DESTINO:")
        canvas_obj.setFont("Helvetica", 12)
        destino = effective_destiny[:35] if len(effective_destiny) > 35 else effective_destiny
        canvas_obj.drawString(content_x, current_y - 0.2*inch, destino)
        current_y -= 0.5*inch
        
        # Agencia de transporte
        canvas_obj.setFont("Helvetica-Bold", 9)
        canvas_obj.drawString(content_x, current_y, "AGENCIA:")
        canvas_obj.setFont("Helvetica", 9)
        agency_name = effective_agency.name if effective_agency else 'Sin asignar'
        agency_display = agency_name[:30] if len(agency_name) > 30 else agency_name
        canvas_obj.drawString(content_x, current_y - 0.15*inch, agency_display)
        current_y -= 0.35*inch
        
        # Número de guía
        canvas_obj.setFont("Helvetica-Bold", 9)
        canvas_obj.drawString(content_x, current_y, "GUÍA:")
        canvas_obj.setFont("Helvetica", 9)
        guide = effective_guide or 'Sin guía'
        guide_display = guide[:30] if len(guide) > 30 else guide
        canvas_obj.drawString(content_x, current_y - 0.15*inch, guide_display)
        current_y -= 0.35*inch
        
        # Cantidad de paquetes
        canvas_obj.setFont("Helvetica-Bold", 9)
        canvas_obj.drawString(content_x, current_y, "PAQUETES:")
        canvas_obj.setFont("Helvetica-Bold", 14)
        packages_count = pull.packages.count()
        canvas_obj.drawString(content_x + 0.8*inch, current_y - 0.05*inch, str(packages_count))
        current_y -= 0.4*inch
        
        # Tamaño de saca
        canvas_obj.setFont("Helvetica-Bold", 8)
        canvas_obj.drawString(content_x, current_y, "TAMAÑO:")
        canvas_obj.setFont("Helvetica", 9)
        canvas_obj.drawString(content_x + 0.6*inch, current_y, pull.get_size_display())
        
        # Generar código QR (solo ID de la saca)
        qr_data = str(pull.id)
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=1,
        )
        qr.add_data(qr_data)
        qr.make(fit=True)
        
        # Convertir QR a imagen
        qr_img = qr.make_image(fill_color="black", back_color="white")
        qr_buffer = io.BytesIO()
        qr_img.save(qr_buffer, format='PNG')
        qr_buffer.seek(0)
        
        # Dibujar QR (centrado en la parte inferior)
        qr_size = 1.3 * inch
        qr_x = x + (label_width - qr_size) / 2
        qr_y = y + 0.3 * inch
        
        # Usar ImageReader para leer el BytesIO
        img_reader = ImageReader(qr_buffer)
        canvas_obj.drawImage(
            img_reader,
            qr_x,
            qr_y,
            width=qr_size,
            height=qr_size,
            preserveAspectRatio=True
        )
        
        # Texto bajo el QR
        canvas_obj.setFont("Helvetica", 7)
        canvas_obj.drawCentredString(
            x + label_width/2,
            y + 0.15*inch,
            f"ID: {str(pull.id)[:8].upper()}"
        )
