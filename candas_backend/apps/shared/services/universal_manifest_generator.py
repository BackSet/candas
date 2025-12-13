"""
Generador universal de manifiestos que funciona con cualquier entidad
usando el patrón Adapter.
"""
import re
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, Spacer, PageBreak, Table, TableStyle
from reportlab.lib.styles import ParagraphStyle
from typing import Any
from apps.shared.services.manifest_template import BaseManifestGenerator
from apps.shared.services.manifest_adapters import ManifestAdapter


class UniversalManifestGenerator:
    """
    Generador genérico que puede crear manifiestos para cualquier entidad
    usando un adapter.
    """
    
    @staticmethod
    def generate_pdf(entity: Any, adapter: ManifestAdapter):
        """
        Genera PDF de manifiesto para cualquier entidad usando un adapter.
        
        Args:
            entity: Cualquier entidad (Batch, Pull, Dispatch, etc.)
            adapter: Instancia de ManifestAdapter que define cómo extraer información
            
        Returns:
            BytesIO con el PDF generado
        """
        doc, buffer = BaseManifestGenerator.create_document()
        elements = []
        
        # Header de empresa (compacto)
        elements.append(BaseManifestGenerator.get_company_header())
        elements.append(Spacer(1, 1*mm))  # Reducido de 3mm a 1mm
        
        # Título
        title_style = BaseManifestGenerator.get_title_style()
        elements.append(Paragraph(adapter.get_title(), title_style))
        elements.append(Spacer(1, 1*mm))  # Reducido de 3mm a 1mm
        
        # Información de la entidad (en tabla compacta de 2 columnas)
        normal_style = BaseManifestGenerator.get_normal_style()
        info_fields = adapter.get_info_fields(entity)
        
        # Crear tabla compacta de información (2 columnas)
        info_table_data = []
        for i in range(0, len(info_fields), 2):
            row = []
            # Usar _wrap_text del BaseManifestGenerator
            row.append(BaseManifestGenerator._wrap_text(
                info_fields[i][0] + ' ' + str(info_fields[i][1]), 
                normal_style
            ))
            if i + 1 < len(info_fields):
                row.append(BaseManifestGenerator._wrap_text(
                    info_fields[i+1][0] + ' ' + str(info_fields[i+1][1]), 
                    normal_style
                ))
            else:
                row.append(BaseManifestGenerator._wrap_text('', normal_style))
            info_table_data.append(row)
        
        if info_table_data:
            info_table = Table(info_table_data, colWidths=[143.5*mm, 143.5*mm])
            info_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 1),
                ('RIGHTPADDING', (0, 0), (-1, -1), 1),
                ('TOPPADDING', (0, 0), (-1, -1), 1),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
            ]))
            elements.append(info_table)
        
        elements.append(Spacer(1, 2*mm))  # Reducido de 5mm a 2mm
        
        # Nota importante (compacta)
        warning_style = ParagraphStyle(
            'WarningStyle',
            parent=normal_style,
            fontSize=6.5,  # Reducido de 8 a 6.5
            textColor=colors.black,
            fontName='Helvetica-Bold',
            spaceAfter=1,  # Reducido de 5 a 1
            spaceBefore=1  # Reducido de 5 a 1
        )
        warning_text = "IMPORTANTE: Priorice la columna NOTAS para confirmar el destino. Si no hay observaciones, guíese por los datos de dirección regular."
        elements.append(Paragraph(warning_text, warning_style))
        elements.append(Spacer(1, 1*mm))  # Reducido de 3mm a 1mm
        
        # Obtener paquetes (pueden venir agrupados o no)
        packages_data = adapter.get_packages(entity)
        
        if not packages_data:
            elements.append(Paragraph("No hay paquetes", normal_style))
        else:
            # Si el adapter retorna lista de dicts (agrupados)
            if isinstance(packages_data, list) and packages_data and isinstance(packages_data[0], dict):
                for group_idx, group_data in enumerate(packages_data, 1):
                    # Solo agregar PageBreak si no es el primer grupo
                    # (el PageBreak se manejará automáticamente si no cabe en la página)
                    if group_idx > 1:
                        # Intentar agregar sin PageBreak primero, solo si es necesario
                        # ReportLab manejará automáticamente los saltos de página
                        pass
                    
                    # Encabezado del grupo (compacto)
                    if group_data.get('group_label'):
                        heading_style = BaseManifestGenerator.get_heading_style()
                        elements.append(Paragraph(group_data['group_label'], heading_style))
                        elements.append(Spacer(1, 0.5*mm))  # Reducido aún más
                    
                    # Información del grupo (en línea compacta)
                    group_info = group_data.get('group_info', [])
                    if group_info:
                        # Mostrar en una sola línea separada por " | "
                        group_info_text = " | ".join([f"{label} {value}" for label, value in group_info])
                        elements.append(Paragraph(group_info_text, normal_style))
                        elements.append(Spacer(1, 0.5*mm))  # Reducido aún más
                    
                    # Tabla de paquetes
                    packages = group_data.get('packages', [])
                    if packages:
                        UniversalManifestGenerator._add_packages_table(elements, packages)
                        # Solo agregar espacio si hay más grupos
                        if group_idx < len(packages_data):
                            elements.append(Spacer(1, 1*mm))  # Espacio mínimo entre grupos
            else:
                # Lista simple de paquetes (sin agrupación)
                UniversalManifestGenerator._add_packages_table(elements, packages_data)
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def _add_packages_table(elements, packages):
        """Agrega tabla de paquetes al documento (optimizada para máximo aprovechamiento)."""
        normal_style = BaseManifestGenerator.get_normal_style()
        data_style = ParagraphStyle(
            'FormalData',
            parent=normal_style,
            fontSize=6,  # Reducido de 7 a 6
            leading=7.2  # Interlineado compacto
        )
        note_style = ParagraphStyle(
            'NoteStyle',
            parent=normal_style,
            fontSize=6,  # Reducido de 7 a 6
            textColor=colors.black,
            fontName='Helvetica-Bold',
            leftIndent=2,  # Reducido de 5 a 2
            rightIndent=2,  # Reducido de 5 a 2
            leading=7.2
        )
        bold_style = ParagraphStyle(
            'FormalBold',
            parent=normal_style,
            fontName='Helvetica-Bold',
            fontSize=7,  # Reducido de 8 a 7
            leading=8
        )
        
        # Encabezados (abreviados para ahorrar espacio)
        headers = ['#', 'Guía', 'Destinatario', 'Dirección', 'Ciudad', 'Prov.', 'Tel.', 'Firma']
        wrapped_headers = [BaseManifestGenerator._wrap_text(h, bold_style) for h in headers]
        table_data = [wrapped_headers]
        
        # Rastrear índices de filas de nota para aplicar SPAN correctamente
        note_row_indices = []
        
        # Datos
        for pkg_idx, package in enumerate(packages, 1):
            pkg_row = [
                BaseManifestGenerator._wrap_text(str(pkg_idx), data_style),
                BaseManifestGenerator._wrap_text(package.guide_number, data_style),
                BaseManifestGenerator._wrap_text(package.name, data_style),
                BaseManifestGenerator._wrap_text(package.address, data_style),
                BaseManifestGenerator._wrap_text(package.city, data_style),
                BaseManifestGenerator._wrap_text(package.province, data_style),
                BaseManifestGenerator._wrap_text(package.phone_number or '-', data_style),
                BaseManifestGenerator._wrap_text('', data_style),
            ]
            table_data.append(pkg_row)
            data_row_idx = len(table_data) - 1  # Índice de la fila de datos que acabamos de agregar
            
            # Fila de nota (solo si existe, más compacta)
            if package.notes:
                # Limpiar las notas ANTES de formatearlas
                # _wrap_text ya limpia, pero asegurémonos de que las notas estén limpias
                clean_notes = str(package.notes).strip() if package.notes else ''
                if clean_notes:
                    note_text = f"NOTA: {clean_notes}"
                    note_cell = BaseManifestGenerator._wrap_text(note_text, note_style)
                    note_row = [note_cell] + [BaseManifestGenerator._wrap_text('', data_style)] * (len(headers) - 2) + [BaseManifestGenerator._wrap_text('', data_style)]
                    table_data.append(note_row)
                    note_row_idx = len(table_data) - 1  # Índice de la fila de nota que acabamos de agregar
                    note_row_indices.append((data_row_idx, note_row_idx))  # Guardar ambos índices
        
        # Anchos de columnas optimizados (aprovechando todo el ancho disponible)
        available_width = 287 * mm  # A4 horizontal (297mm) - márgenes (5mm * 2) = 287mm
        col_widths = [
            12*mm,   # # (reducido de 15mm)
            32*mm,   # Guía (reducido de 35mm)
            38*mm,   # Destinatario (reducido de 40mm)
            48*mm,   # Dirección (reducido de 50mm)
            28*mm,   # Ciudad (reducido de 30mm)
            25*mm,   # Prov. (reducido de 30mm)
            28*mm,   # Tel. (reducido de 30mm)
            36*mm    # Firma (reducido de 47mm)
        ]
        
        # Crear tabla
        table = Table(table_data, colWidths=col_widths)
        
        # Estilo optimizado (padding mínimo, bordes finos)
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),  # Reducido de 8 a 7
            ('BOTTOMPADDING', (0, 0), (-1, 0), 1),  # Reducido de 3 a 1
            ('TOPPADDING', (0, 0), (-1, 0), 1),  # Reducido de 3 a 1
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),  # Reducido de 1.5 a 1
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 6),  # Reducido de 7 a 6
            ('GRID', (0, 0), (-1, -1), 0.3, colors.black),  # Reducido de 0.5 a 0.3
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Cambiado de MIDDLE a TOP para mejor aprovechamiento
            ('LEFTPADDING', (0, 0), (-1, -1), 1),  # Reducido de 2 a 1
            ('RIGHTPADDING', (0, 0), (-1, -1), 1),  # Reducido de 2 a 1
            ('TOPPADDING', (0, 1), (-1, -1), 1),  # Reducido de 2 a 1
            ('BOTTOMPADDING', (0, 1), (-1, -1), 1),  # Reducido de 2 a 1
        ]
        
        # Aplicar estilo a filas de nota (usando índices rastreados)
        firma_col_idx = len(headers) - 1
        for data_row_idx, note_row_idx in note_row_indices:
            # Nota ocupa todas las columnas excepto Firma (desde columna 0 hasta penúltima)
            table_style.append(('SPAN', (0, note_row_idx), (-2, note_row_idx)))
            table_style.append(('BACKGROUND', (0, note_row_idx), (-2, note_row_idx), colors.HexColor('#f5f5f5')))
            table_style.append(('FONTNAME', (0, note_row_idx), (-2, note_row_idx), 'Helvetica-Bold'))
            table_style.append(('TOPPADDING', (0, note_row_idx), (-2, note_row_idx), 0.5))  # Padding mínimo
            table_style.append(('BOTTOMPADDING', (0, note_row_idx), (-2, note_row_idx), 0.5))  # Padding mínimo
            # Unir la columna de Firma verticalmente (fila del paquete + fila de nota)
            table_style.append(('SPAN', (firma_col_idx, data_row_idx), (firma_col_idx, note_row_idx)))
            # Eliminar bordes horizontales entre filas para unificar visualmente
            table_style.append(('LINEBELOW', (0, data_row_idx), (-2, data_row_idx), 0, colors.white))
            table_style.append(('LINEABOVE', (0, note_row_idx), (-2, note_row_idx), 0, colors.white))
        
        table.setStyle(TableStyle(table_style))
        elements.append(table)
    
    @staticmethod
    def _clean_text_for_excel(text):
        """
        Limpia texto para Excel eliminando saltos de línea y caracteres especiales.
        
        Args:
            text: Texto a limpiar
            
        Returns:
            Texto limpio o '-' si está vacío
        """
        if not text:
            return '-'
        text_str = str(text)
        # Limpiar saltos de línea
        text_str = text_str.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
        # Eliminar caracteres de control
        text_str = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F]', '', text_str)
        # Eliminar caracteres Unicode problemáticos
        text_str = re.sub(r'[\u2000-\u206F]', '', text_str)
        text_str = re.sub(r'[\u2028\u2029]', ' ', text_str)
        text_str = text_str.replace('\uFEFF', '').replace('\u200B', '').replace('\u200C', '').replace('\u200D', '').replace('\u2060', '')
        # Normalizar espacios
        text_str = re.sub(r' +', ' ', text_str).strip()
        return text_str if text_str else '-'
    
    @staticmethod
    def generate_excel(entity: Any, adapter: ManifestAdapter):
        """Genera Excel de manifiesto."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from datetime import datetime
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Manifiesto"
        
        # Estilos optimizados
        header_font = Font(bold=True, size=10)  # Reducido de 12 a 10
        normal_font = Font(size=8)  # Reducido de 10 a 8
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título (compacto)
        ws.merge_cells('A1:H1')
        ws['A1'] = adapter.get_title()
        ws['A1'].font = Font(bold=True, size=12)  # Reducido de 16 a 12
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 18  # Reducido de 25 a 18
        
        # Información (compacta, 2 columnas)
        row = 2  # Reducido de 3 a 2
        info_fields = adapter.get_info_fields(entity)
        # Mostrar en 2 columnas para ahorrar espacio
        for i in range(0, len(info_fields), 2):
            ws[f'A{row}'] = info_fields[i][0] + ' ' + str(info_fields[i][1])
            ws[f'A{row}'].font = normal_font
            if i + 1 < len(info_fields):
                ws[f'E{row}'] = info_fields[i+1][0] + ' ' + str(info_fields[i+1][1])
                ws[f'E{row}'].font = normal_font
            row += 1
        
        # Paquetes
        packages_data = adapter.get_packages(entity)
        
        if not packages_data:
            row += 1
            ws[f'A{row}'] = 'No hay paquetes'
        else:
            # Si viene agrupado
            if isinstance(packages_data, list) and packages_data and isinstance(packages_data[0], dict):
                for group_data in packages_data:
                    row += 1  # Reducido de 2 a 1
                    if group_data.get('group_label'):
                        ws[f'A{row}'] = group_data['group_label']
                        ws.merge_cells(f'A{row}:H{row}')
                        ws[f'A{row}'].font = header_font
                        ws[f'A{row}'].alignment = Alignment(horizontal='center')
                        ws.row_dimensions[row].height = 15  # Altura compacta
                    
                    group_info = group_data.get('group_info', [])
                    if group_info:
                        row += 1
                        # Mostrar en una sola línea
                        group_info_text = " | ".join([f"{label} {value}" for label, value in group_info])
                        ws[f'A{row}'] = group_info_text
                        ws[f'A{row}'].font = normal_font
                        ws.row_dimensions[row].height = 12  # Altura compacta
                    
                    packages = group_data.get('packages', [])
                    if packages:
                        row += 1
                        headers = ['#', 'Guía', 'Destinatario', 'Dirección', 'Ciudad', 'Provincia', 'Teléfono', 'Nota']
                        for col, header in enumerate(headers, 1):
                            cell = ws.cell(row, col, header)
                            cell.font = header_font
                            cell.border = border
                        row += 1
                        
                        for pkg_idx, package in enumerate(packages, 1):
                            ws.cell(row, 1, pkg_idx).border = border
                            ws.cell(row, 2, UniversalManifestGenerator._clean_text_for_excel(package.guide_number)).border = border
                            ws.cell(row, 3, UniversalManifestGenerator._clean_text_for_excel(package.name)).border = border
                            ws.cell(row, 4, UniversalManifestGenerator._clean_text_for_excel(package.address)).border = border
                            ws.cell(row, 5, UniversalManifestGenerator._clean_text_for_excel(package.city)).border = border
                            ws.cell(row, 6, UniversalManifestGenerator._clean_text_for_excel(package.province)).border = border
                            ws.cell(row, 7, UniversalManifestGenerator._clean_text_for_excel(package.phone_number)).border = border
                            note_cell = ws.cell(row, 8, UniversalManifestGenerator._clean_text_for_excel(package.notes))
                            note_cell.border = border
                            note_cell.font = Font(bold=True)
                            note_cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                            row += 1
            else:
                # Lista simple
                row += 1
                headers = ['#', 'Guía', 'Destinatario', 'Dirección', 'Ciudad', 'Provincia', 'Teléfono', 'Nota']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row, col, header)
                    cell.font = header_font
                    cell.border = border
                row += 1
                
                for pkg_idx, package in enumerate(packages_data, 1):
                    ws.cell(row, 1, pkg_idx).border = border
                    ws.cell(row, 2, UniversalManifestGenerator._clean_text_for_excel(package.guide_number)).border = border
                    ws.cell(row, 3, UniversalManifestGenerator._clean_text_for_excel(package.name)).border = border
                    ws.cell(row, 4, UniversalManifestGenerator._clean_text_for_excel(package.address)).border = border
                    ws.cell(row, 5, UniversalManifestGenerator._clean_text_for_excel(package.city)).border = border
                    ws.cell(row, 6, UniversalManifestGenerator._clean_text_for_excel(package.province)).border = border
                    ws.cell(row, 7, UniversalManifestGenerator._clean_text_for_excel(package.phone_number)).border = border
                    note_cell = ws.cell(row, 8, UniversalManifestGenerator._clean_text_for_excel(package.notes))
                    note_cell.border = border
                    note_cell.font = Font(bold=True)
                    note_cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                    row += 1
        
        # Ajustar anchos (optimizados)
        ws.column_dimensions['A'].width = 6  # Reducido de 8
        ws.column_dimensions['B'].width = 16  # Reducido de 18
        ws.column_dimensions['C'].width = 20  # Reducido de 22
        ws.column_dimensions['D'].width = 26  # Reducido de 28
        ws.column_dimensions['E'].width = 16  # Reducido de 18
        ws.column_dimensions['F'].width = 16  # Reducido de 18
        ws.column_dimensions['G'].width = 13  # Reducido de 15
        ws.column_dimensions['H'].width = 28  # Reducido de 30
        
        # Ajustar altura de filas de datos para compactar
        for row_num in range(1, ws.max_row + 1):
            if ws.row_dimensions[row_num].height is None:
                ws.row_dimensions[row_num].height = 12  # Altura compacta por defecto
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

